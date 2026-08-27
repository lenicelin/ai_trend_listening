#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AI Trend Listening - Stage 2: Weekly Newsletter Curator Engine
Reads Stage 1's database (data/stage1_ai_news.json), dynamically reads active theme from
data/weekly_newsletter_theme.xlsx (managed by user), evaluates articles based on 4-tier criteria,
selects top articles matching the active weekly theme, generates curator rationales,
and exports to data/stage2_curated_news.xlsx and data/stage2_curated_report.md.
"""

import sys
import os
import json
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Force UTF-8 encoding for stdout on Windows
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding='utf-8', line_buffering=True)
    except Exception:
        pass

# Ensure environment variables are loaded from .env if present
if os.path.exists(".env"):
    try:
        with open(".env", "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, v = line.split("=", 1)
                    os.environ[k.strip()] = v.strip()
    except Exception:
        pass

# Use fast native urllib REST API for Gemini calls
HAS_GENAI = False

DEFAULT_THEME = "AI 重塑自動化：從單機控制到協同智慧製造"
DEFAULT_ISSUE = "Vol. 2026 Issue #4"
DEFAULT_DOMAINS = "自動化、工業自動化、自主機器人"

_DOMAIN_EXPANSION_CACHE = {}
_LAST_LLM_CALL_TIME = 0

def enforce_rate_limit_delay(min_interval: float = 4.0):
    global _LAST_LLM_CALL_TIME
    import time
    now = time.time()
    elapsed = now - _LAST_LLM_CALL_TIME
    if elapsed < min_interval:
        sleep_dur = min_interval - elapsed
        time.sleep(sleep_dur)
    _LAST_LLM_CALL_TIME = time.time()

def call_llm_api(prompt: str, timeout: int = 25) -> str:
    """
    Unified LLM API Invoker calling Gemini API with rate-limiting buffer.
    """
    enforce_rate_limit_delay(0.8)

    gemini_key = os.environ.get("GEMINI_API_KEY")
    if not gemini_key:
        raise ValueError("GEMINI_API_KEY not found in environment variables.")

    import urllib.request
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-flash-latest:generateContent?key={gemini_key}"
    payload = {"contents": [{"parts": [{"text": prompt}]}]}
    req = urllib.request.Request(url, data=json.dumps(payload).encode('utf-8'), headers={'Content-Type': 'application/json'}, method='POST')
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        res_data = json.loads(resp.read().decode('utf-8'))
        return res_data['candidates'][0]['content']['parts'][0]['text']

def expand_domain_keywords_via_llm(domain_name: str) -> list:
    """Tier 1: LLM Dynamic Keyword & Domain Expansion via Gemini API."""
    if domain_name in _DOMAIN_EXPANSION_CACHE:
        return _DOMAIN_EXPANSION_CACHE[domain_name]

    if not os.environ.get("GEMINI_API_KEY"):
        return []

    prompt = f"""你是一位專業科技與 AI 新聞編輯。請針對新聞搜尋主題/領域：【{domain_name}】
動態擴充出相關的中文與英文關鍵字、專業術語、知名品牌廠商、核心技術與英文縮寫。
格式要求：請僅輸出一個 JSON 格式的字串陣列 (Array of Strings)，不要包含 Markdown 標記，不要附加任何說明文字。
範例輸出：
["電動車", "EV", "Tesla", "特斯拉", "BYD", "比亞迪", "Battery", "BMS", "Autonomous Driving", "自動駕駛", "ADAS", "車用晶片"]
"""
    try:
        raw_response_text = call_llm_api(prompt, timeout=6)
        text = raw_response_text.strip().replace("```json", "").replace("```", "").strip()
        kws = json.loads(text)
        if isinstance(kws, list):
            kws_clean = [str(k).strip().lower() for k in kws if str(k).strip()]
            _DOMAIN_EXPANSION_CACHE[domain_name] = kws_clean
            print(f"🤖 [Tier 1 LLM Expansion] 【{domain_name}】 -> 成功動態擴充 {len(kws_clean)} 個關鍵字 (例如: {kws_clean[:5]})")
            return kws_clean
    except Exception as e:
        print(f"⚠️ [Tier 1 LLM Notice] 關鍵字擴充降級跳過 ({e})")
    return []

def get_active_newsletter_theme():
    """Dynamically fetch the active weekly theme & curation goal description from DB or Excel."""
    excel_path = "data/weekly_newsletter_theme.xlsx"
    if not os.path.exists(excel_path):
        print(f"ℹ️ {excel_path} not found. Using default theme.")
        return DEFAULT_THEME, DEFAULT_ISSUE, DEFAULT_DOMAINS, ""
        
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb["每周電子報主題設定"]
        for row in range(5, ws.max_row + 1):
            row_vals = [str(ws.cell(row=row, column=c).value or "").strip() for c in range(1, max(ws.max_column + 1, 10))]
            if any("啟用中" in v or "Active" in v for v in row_vals):
                issue = row_vals[0] or DEFAULT_ISSUE
                theme = row_vals[2] or DEFAULT_THEME
                domains_raw = row_vals[3] or ""
                goal_desc = row_vals[4] if len(row_vals) > 4 and "啟用中" not in row_vals[4] else ""
                
                if not domains_raw or domains_raw == "None":
                    raw_kws = [w.strip() for w in re.split(r'[,:\s/]+', theme) if len(w.strip()) >= 2 and w.strip().lower() not in ["ai", "從", "到", "的", "與", "重塑"]]
                    domains = "、".join(raw_kws) if raw_kws else "AI 應用、產業轉型"
                else:
                    domains = domains_raw

                print(f"🎯 [Stage 2 Theme Manager] Successfully loaded active theme from Excel:\n   期數: {issue}\n   主題: {theme}\n   重點領域: {domains}\n   選題目標描述: {goal_desc[:60]}...")
                return theme, issue, domains, goal_desc
    except Exception as e:
        print(f"⚠️ Warning: Could not read theme Excel file ({e}). Fallback to default theme.")
        
    return DEFAULT_THEME, DEFAULT_ISSUE, DEFAULT_DOMAINS, ""

# Extended Domain Knowledge Base mapping active topics to comprehensive matching keywords
DOMAIN_EXPANSION = {
    "自動化": ["自動化", "automation", "協同", "單機控制", "手臂", "機械手臂", "機器人", "robot", "cobot", "plc", "agv", "amr", "scada", "ot", "工業4.0", "驅動", "控制", "流程自動化"],
    "工業自動化": ["工業自動化", "工業4.0", "智慧製造", "製造", "工廠", "生產線", "良率", "設備", "機台", "半導體", "晶片", "晶圓", "機房", "industrial", "factory", "manufacturing", "semiconductor", "yield"],
    "自主機器人": ["自主機器人", "機器人", "robot", "robotics", "cobot", "具身智能", "embodied ai", "人形機器人", "agv", "amr", "自駕", "移動機器人", "巡檢"],
    "電動車": ["電動車", "電池", "智慧駕駛", "車聯網", "自駕", "車用", "特斯拉", "tesla", "ev", "autonomous", "battery", "byd", "鴻海"],
    "電池管理": ["電池", "電池管理", "續航", "充電", "儲能", "battery", "bms", "充電樁"],
    "智慧駕駛": ["智慧駕駛", "自駕", "駕駛", "fsd", "adas", "車載", "車聯網", "autonomous driving"],
    "研發": ["研發", "r&d", "research", "lab", "實驗室", "開發流程", "工程團隊", "試驗", "演算法開發", "科學家", "論文"],
    "材料": ["材料", "material", "半導體", "晶片", "矽", "電池材料", "化學", "新材料", "封裝", "晶圓", "奈米"],
    "產品生命週期": ["產品生命週期", "設計週期", "lifecycle", "eda", "cad", "品質控管", "規格拆解", "軟體開發", "需求拆解", "程式生成", "測試與除錯", "晶片設計", "產品設計"],
    "策略決策": ["策略", "決策", "商業智慧", "洞察", "轉型", "策略佈局", "領導力", "企劃", "strategy", "decision"],
    "數據洞察": ["數據洞察", "數據分析", "商業智慧", "bi", "預測", "分析模型", "data insight", "analytics"],
    "風險預警": ["風險預警", "風險控管", "資安", "合規", "監管", "安全", "隱私", "預警", "risk", "security", "compliance"],
    "供應鏈": ["供應鏈", "物流", "倉儲", "運籌", "庫存", "航運", "貨運", "採購", "關稅", "supply chain", "logistics", "warehouse", "inventory", "shipping"]
}

GENERIC_NOISE_TERMS = {
    "ai", "nvidia", "amd", "intel", "google", "microsoft", "apple", "amazon", "meta",
    "cloud", "software", "hardware", "chip", "chips", "model", "llm", "agent",
    "ride", "drive", "data", "r&d", "tech", "news", "report", "system",
    "從", "到", "的", "與", "重塑", "加速", "驅動", "革命", "應用"
}

# Precision Scoring Dictionaries (Evaluated in Stage 2 to Maximize Precision)
USE_CASE_TERMS = [
    "客服", "招募", "履歷", "供應鏈", "推薦", "程式碼", "審查", "對話", "預測", "維護", "自動化", "轉型", "案例", "防護", "合規", "機房", "良率", "行銷", "業務", "零售", "電商", "資安", "數據", "倉儲", "物流", "晶片", "伺服器", "產線", "製造", "工廠", "營運", "工作流", "治理", "人才", "機密", "保護", "體驗", "招聘", "服務", "平台", "範例", "研發", "科學", "材料", "電動車", "電池", "駕駛", "自駕", "車聯網", "車用", "車輛", "智慧交通",
    "customer service", "recruiting", "supply chain", "recommendation", "code review", "prediction", "maintenance", "automation", "compliance", "fraud", "marketing", "sales", "retail", "logistics", "warehouse", "chip", "factory", "workflow", "hiring", "data", "research", "science", "ev", "battery", "autonomous", "driving", "robotics", "vision", "audio", "agent", "agentic", "deployment", "training", "generation", "multimodal", "speech", "forecast", "forecasting", "healthcare", "drug", "weather", "code", "coding", "reasoning", "search", "analysis", "embedding", "model", "models"
]

TECH_TERMS = [
    "llama", "claude", "gemini", "gpt", "deepseek", "rag", "agent", "agentic", "vector", "langchain", "autogen", "fine-tuning", "vertex ai", "copilot", "bedrock", "transformer", "模型", "大模型", "架構", "多模態", "api", "即時防護", "雲地混合", "演算法", "軟體", "服務", "平台", "技術", "系統", "硬體", "晶片", "加速器", "處理器", "伺服器", "數據中心", "雲端", "alphafold",
    "model", "system", "platform", "software", "hardware", "cloud", "ai", "llm"
]

IMPACT_TERMS = [
    "%", "成", "秒", "分", "倍", "節省", "縮短", "提升", "降低", "方案", "步驟", "指引", "框架", "處置", "解決方案", "風險", "挑戰", "效益", "成效", "成果", "應用", "突破", "轉型", "新制", "戰略", "法規", "裁判", "判決", "規範", "範例", "新局", "美元", "億", "萬", "roi", "成本", "營收",
    "percent", "reduced", "improved", "faster", "decreased", "solution", "framework", "guideline", "step", "challenge", "impact", "policy", "rule", "strategy", "introducing", "launching", "open", "breakthrough", "accelerating", "new", "performance", "benchmark", "eval", "evaluation", "sota", "state-of-the-art", "release", "announcing"
]

# Code Review Item 5: Dictionary-Based Department Tag Mapping
DEPARTMENT_RULES = [
    {
        "dept": "財務",
        "title_keywords": ["金融", "支付", "交易", "財務", "資金"],
        "text_keywords": ["金融", "支付", "交易", "經手支付"]
    },
    {
        "dept": "供應鏈",
        "title_keywords": ["walmart", "永續零售", "食品浪費"],
        "text_keywords": ["供應鏈", "物流", "減碳"]
    },
    {
        "dept": "資安",
        "title_keywords": ["叡揚", "資安閉環", "單點風險", "資安", "安全與防禦", "漏洞"],
        "text_keywords": ["微隔離", "微分段", "資安閉環", "漏洞", "駭客"]
    },
    {
        "dept": "製造",
        "title_keywords": ["自動化展", "台達", "跨廠部署", "智造", "產線"],
        "text_keywords": ["跨廠", "數位雙生", "噴塗膠", "智能製造"]
    },
    {
        "dept": "高階治理",
        "title_keywords": ["連署", "放慢", "頂尖ai", "公開信", "治理"],
        "text_keywords": ["連署", "放慢頂尖", "公開信"]
    },
    {
        "dept": "研發",
        "title_keywords": ["claude code", "gemini cli", "codex", "程式", "開源", "r&d", "研發"],
        "text_keywords": ["claude code", "codex", "r&d"]
    }
]

def infer_department_tag(title: str, full_text: str = "") -> str:
    """Dictionary-based department tag inference (Refactored for maintainability)."""
    t = title.lower()
    text = (title + " " + full_text).lower()

    for rule in DEPARTMENT_RULES:
        if any(k in t for k in rule["title_keywords"]):
            return rule["dept"]

    for rule in DEPARTMENT_RULES:
        if any(k in text for k in rule["text_keywords"]):
            return rule["dept"]

    return "高階治理"

def check_hard_constraint(full_text: str, domain_keyword_map: dict, active_theme: str = ""):
    """
    Code Review Item 1 & 2: Single Source of Truth Hard Constraint & Domain Matcher.
    Checks regex patterns AND matches keywords from domain_keyword_map in one unified pass.
    Returns (passed: bool, matched_domains: list, matched_kws: list).
    """
    all_text = f"{active_theme} {' '.join(domain_keyword_map.keys())}".lower()
    
    # 1. EV Domain Regex Check
    if any(k in all_text for k in ["電動車", "車用", "智慧駕駛", "自駕", "車載", "adas", "bms", "ev", "汽車"]):
        ev_pattern = r'(電動車|智慧電動車|ai汽車|車用|車載|自駕|自動駕駛|智慧駕駛|電池管理|電池壽命|電池熱管理|特斯拉|車用晶片|車用ai|車用電子|車載運算|光達|\b(ev|evs|bms|adas|lidar|tesla|nvidia drive|qualcomm ride|snapdragon ride|autonomous driving|autonomous vehicle|autonomous car)\b)'
        if not re.search(ev_pattern, full_text, re.IGNORECASE):
            return False, [], []

    # 2. Smart Manufacturing Regex Check
    if any(k in all_text for k in ["自動化", "智慧製造", "工業", "機器人", "agv", "amr", "cobot"]):
        factory_pattern = r'(自動化|智慧製造|工業自動化|機械手臂|機器人|具身智能|智慧工廠|\b(cobot|plc|agv|amr|scada|ot|embodied ai)\b)'
        if not re.search(factory_pattern, full_text, re.IGNORECASE):
            return False, [], []

    matched_domains = []
    matched_kws = []

    # 3. Single Pass Match against domain_keyword_map (Single Source of Truth)
    for dom_name, kws in domain_keyword_map.items():
        found = [kw for kw in kws if kw in full_text]
        if found:
            matched_domains.append(dom_name)
            matched_kws.extend(found)

    if matched_domains:
        return True, matched_domains, matched_kws

    # Fallback: check if domain names themselves appear in text
    for dom_name in domain_keyword_map.keys():
        if dom_name.lower() in full_text:
            matched_domains.append(dom_name)

    if matched_domains:
        return True, matched_domains, matched_domains

    return False, [], []

def build_active_theme_matcher(active_theme, focus_domains, goal_desc=""):
    """Dynamically construct focus domain keyword groups with Tier 1 LLM Expansion fallback."""
    raw_domain_tags = [d.strip() for d in re.split(r'[,:\s/、]+', focus_domains) if len(d.strip()) >= 2]
    theme_terms = [t.strip() for t in re.split(r'[,:\s/、：-]+', active_theme) if len(t.strip()) >= 2 and t.strip().lower() not in ["ai", "從", "到", "的", "與", "重塑", "加速", "驅動"]]
    all_target_domains = list(dict.fromkeys(raw_domain_tags + [t for t in theme_terms if not any(t in d for d in raw_domain_tags)]))
    
    domain_keyword_map = {}
    import time
    for dom in all_target_domains:
        kws = set()
        kws.add(dom.lower())
        
        # 1. Static Dictionary Merge
        static_matched = False
        for k_key, k_list in DOMAIN_EXPANSION.items():
            if k_key in dom or dom in k_key:
                for item in k_list:
                    kws.add(item.lower())
                static_matched = True
                    
        # 2. Tier 1: LLM Dynamic Expansion only if static dictionary had no matches
        if not static_matched:
            llm_kws = expand_domain_keywords_via_llm(dom)
            for item in llm_kws:
                kws.add(item.lower())
            
        domain_keyword_map[dom] = list(kws)
        time.sleep(0.5)
        
    return all_target_domains, domain_keyword_map

def evaluate_article(art, active_theme, focus_domains, target_domains, domain_keyword_map, goal_desc=""):
    """
    Code Review Item 1, 2, 3: Unified article evaluator.
    Directly consumes matched_domains & matched_kws from check_hard_constraint and returns score & tags.
    """
    title = art["title"]
    desc = art.get("description", "")
    source = art.get("source", "")
    full_text = f"{title} {desc} {source}".lower()
    
    passed, matched_domains, matched_kws = check_hard_constraint(full_text, domain_keyword_map, active_theme)
    if not passed or not matched_domains:
        return 0, []

    # 1. Theme score (0-55 pts)
    title_text = title.lower()
    title_matches = sum(1 for dom in matched_domains if dom.lower() in title_text)
    score_theme = min(title_matches * 18 + len(matched_domains) * 12 + len(matched_kws) * 4, 55)

    # 2. Specific Use Case score (0-15 pts)
    use_case_matches = sum(1 for uc in USE_CASE_TERMS if uc in full_text)
    score_use_case = min(use_case_matches * 3, 15)

    # 3. Technical & Architectural Details score (0-15 pts)
    tech_matches = sum(1 for tc in TECH_TERMS if tc in full_text)
    score_tech = min(tech_matches * 3, 15)

    # 4. Quantified Impact & ROI score (0-10 pts)
    impact_matches = sum(1 for im in IMPACT_TERMS if im in full_text)
    score_impact = min(impact_matches * 2, 10)

    # 5. Media Authority & Timeliness score (0-5 pts)
    auth_media = [
        "iThome", "TechCrunch", "Wired", "紐約時報", "經理人", "科技新報", "鉅亨網", "UDN", "風傳媒", 
        "104", "數位時代", "天下雜誌", "商業周刊", "DIGITIMES", "經濟日報", "Business Wire", "NVIDIA", "Google", "Ars Technica", "VentureBeat"
    ]
    score_auth = 5 if any(m in source for m in auth_media) else 2

    total_score = score_theme + score_use_case + score_tech + score_impact + score_auth
    
    dept_tag = infer_department_tag(title, full_text)
    functional_tags = [f"{dept_tag} / {dom}" for dom in matched_domains[:2]]
    if not functional_tags:
        functional_tags = [f"{dept_tag} / {focus_domains.split('、')[0]}"]

    return total_score, functional_tags

def evaluate_candidates_via_llm(candidates, active_theme, focus_domains, goal_desc=""):
    """Tier 2: LLM Deep Semantic Scoring & Curator Rationale Generation via Gemini API."""
    if not os.environ.get("GEMINI_API_KEY") or not candidates:
        return candidates

    eval_candidates = candidates[:10]
    print(f"🤖 [Tier 2 LLM Evaluation] Submitting top {len(eval_candidates)} candidates to LLM for deep semantic curation...")

    prompt_items = []
    for i, c in enumerate(eval_candidates):
        prompt_items.append({
            "id": i,
            "title": c["title"],
            "source": c["source"],
            "description": c.get("description", "")[:200]
        })

    prompt = f"""你是一位權威科技與 AI 電子報總主編。本期電子報主題為：
《{active_theme}》
重點關注領域：{focus_domains}
選題目標：{goal_desc}

請幫我分析與評估以下 {len(eval_candidates)} 篇新聞候選名單，並進行「深度語意精準打分 (Maximize Precision)」與「專業選文理由撰寫」：

候選新聞列表 (JSON):
{json.dumps(prompt_items, ensure_ascii=False)}

評分與審核要求（請務必綜合評估：1. 主題契合度、2. 具體應用場景、3. 技術架構細節、4. 量化 ROI 與效益數據，並嚴格執行評分階梯拉開與相對排序）：
1. score (整數 0-100 分)：
   - Top-Tier (極具代表性、最符合主題，且同時具備具體落地案例、技術架構細節或量化成效數據的前 2-3 篇)：給予 90-98 分
   - Mid-Tier (深度契合主題、具重要實務參考價值的用例/技術報導前 4-8 篇)：給予 80-89 分
   - Standard-Tier (符合主題但屬一般性動態或缺乏具體量化數據的前 9-15 篇)：給予 70-79 分
   - Low-Tier / Margin-Tier (主題相關度較低或內容較空泛的文章)：給予 60-69 分或更低
   * 注意：請確保每篇文章的分數具有明顯鑑別度與階梯差距（相鄰排名建議相差 1-3 分），絕對禁止多篇打相同高分（如通通 96 或 98 分）。

2. functional_tags (字串)：為該文章指定職能對應標籤（格式如：「智慧製造 / 自主機器人」或「企業轉型 / 策略決策」）。
3. rationale (字串)：撰寫 2-3 句極具商業洞察力的「選入理由與推薦切入點」，說明此報導如何對企業主管產生決策價值。格式範例：「對應領域：自主機器人。\n切入點：【自主機器人視角】聚焦智慧工廠無人搬運系統落地實例，解析如何提升生產良率。」

請嚴格僅輸出一個 JSON 陣列，每個元素包含: "id", "score", "functional_tags", "rationale"。
不要包含 Markdown 標籤，不要輸出任何額外文字。
"""
    for attempt in range(2):
        try:
            if attempt > 0:
                import time
                time.sleep(2)
            raw_response_text = call_llm_api(prompt, timeout=90)
            text = raw_response_text.strip().replace("```json", "").replace("```", "").strip()
            eval_results = json.loads(text)

            eval_map = {item["id"]: item for item in eval_results if isinstance(item, dict) and "id" in item}
            updated_candidates = []
            for idx, cand in enumerate(candidates):
                if idx in eval_map:
                    res_item = eval_map[idx]
                    cand["score"] = int(res_item.get("score", cand["score"]))
                    cand["functional_tags"] = str(res_item.get("functional_tags", cand["functional_tags"]))
                    cand["rationale"] = str(res_item.get("rationale", cand.get("rationale", "")))
                updated_candidates.append(cand)

            print(f"✨ [Tier 2 LLM Evaluation SUCCESS] Re-scored & generated rationales for {len(updated_candidates)} candidates.")
            return updated_candidates
        except Exception as e:
            if attempt == 1:
                print(f"⚠️ [Tier 2 LLM Notice] LLM 語意打分降級，使用 Tier 1 規則評分結果 ({e})")
    return candidates

def synthesize_weekly_express_theme(articles):
    """Auto-synthesizes dominant weekly trend title for Express Mode from Stage 1 news via Gemini API."""
    recent_titles = [f"• {a['title']} ({a.get('source', '')})" for a in articles[:15]]
    titles_block = "\n".join(recent_titles)
    
    prompt = f"""你是一位權威科技與 AI 電子報總主編。請閱讀以下本週最新採集的科技與 AI 新聞標題列表：

{titles_block}

請幫我提煉出本週科技與 AI 界最主要談論的 1 個核心趨勢議題，並合成為一個精煉、吸引企業主管閱讀的「每週快報專題標題」(15字以內)。
範例：
- AI Agent 權限治理與雲端算力大廠角力
- 人機協作重塑企業生產力與資安防禦
- 具身智能與半導體供應鏈新突破

請僅輸出標題文字本身，不要包含任何標籤、說明或引號。"""

    if os.environ.get("GEMINI_API_KEY"):
        try:
            synthesized = call_llm_api(prompt, timeout=12).strip()
            clean_title = re.sub(r'^["\'「」【】]+|["\'「」【】]+$', '', synthesized).strip()
            if clean_title and len(clean_title) <= 30:
                print(f"🤖 [LLM Express Theme Synthesis] Synthesized weekly trend theme: 【{clean_title}】")
                return f"每週 AI 趨勢快報：{clean_title}"
        except Exception as e:
            print(f"⚠️ [Express Theme Synthesis Notice] LLM synthesis fallback: {e}")

    # Fallback heuristic title based on keywords in titles
    all_text = " ".join([a["title"] for a in articles[:10]])
    if "Agent" in all_text or "代理" in all_text:
        return "每週 AI 趨勢快報：AI Agent 自動化與資安防禦前瞻"
    elif "算力" in all_text or "晶片" in all_text or "Azure" in all_text:
        return "每週 AI 趨勢快報：雲端算力基礎設施與大廠生態布局"
    else:
        return "每週 AI 趨勢快報：前瞻 AI 應用與企業數位轉型焦點"


def prepare_curated_export_dto(selected_articles, qualified_threshold=80):
    """
    Code Review Item 6: Unified Data Transfer Object (DTO) Renderer.
    Standardizes output fields across Excel & Markdown exporters to avoid duplicated formatting code.
    """
    if not selected_articles:
        return [{
            "rank": 1,
            "score_str": "0 分",
            "functional_tags": "無符合門檻案例",
            "title": "⚠️ 本期符合主題之高品質新聞不足",
            "pub_date": "-",
            "link": "-",
            "source": "-",
            "rationale": f"堅持「零湊數原則」：本期無文章達到最低合格門檻 (Score >= {qualified_threshold} 分)。系統不會為了湊滿篇數而降低標準或收錄低相關新聞。",
            "description": "-"
        }]

    dto_list = []
    for idx, item in enumerate(selected_articles, 1):
        dto_list.append({
            "rank": idx,
            "score_str": f"{item['score']} 分",
            "functional_tags": item.get("functional_tags", "當週重磅新聞 / 科技趨勢"),
            "title": item["title"],
            "pub_date": item.get("pub_date", "-"),
            "link": item.get("link", "-"),
            "source": item.get("source", "-"),
            "rationale": item.get("rationale", f"對應主題《{item.get('title', '')[:20]}》核心技術落地與實務應用。"),
            "description": item.get("description", item.get("summary", ""))
        })
    return dto_list


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Stage 2 Curator Engine")
    parser.add_argument("--mode", choices=["theme", "dynamic"], default=os.environ.get("PIPELINE_MODE", "theme"),
                        help="Curation mode: 'theme' (特定預設主題 20 篇) or 'dynamic' (全網無預設主題探索)")
    parser.add_argument("--since", default=os.environ.get("START_DATE", "2026-08-01"),
                        help="Filter articles published on or after date (default: '2026-08-01')")
    args, unknown = parser.parse_known_args()
    pipeline_mode = args.mode.lower()
    since_date = args.since.strip()

    json_path = "data/stage1_ai_news.json"
    articles = []
    try:
        from db_manager import get_all_articles
        articles = get_all_articles()
        if articles:
            print(f"📦 Loaded {len(articles)} articles from PostgreSQL Database.")
    except Exception:
        pass

    if not articles:
        if not os.path.exists(json_path) and os.path.exists("data/agent_a_ai_news.json"):
            json_path = "data/agent_a_ai_news.json"

        if os.path.exists(json_path):
            with open(json_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            articles = data.get("articles", [])
            print(f"📦 Loaded {len(articles)} articles from Stage 1 accumulated database ({os.path.basename(json_path)}).")

    if not articles:
        print(f"❌ Error: Cannot find Stage 1 database at {json_path}")
        return

    if since_date:
        filtered_articles = [a for a in articles if str(a.get("pub_date", "")).strip() >= since_date]
        print(f"📅 [Date Filter] Filtered down from {len(articles)} to {len(filtered_articles)} articles published on or after {since_date}.")
        articles = filtered_articles

    theme_from_excel, issue_tag_excel, focus_domains_excel, goal_desc_excel = get_active_newsletter_theme()
    is_excel_dynamic = any(k in theme_from_excel.lower() for k in ["動態", "無預設", "全網", "auto", "dynamic"])

    if pipeline_mode == "dynamic" or is_excel_dynamic:
        pipeline_mode = "dynamic"
        issue_tag = issue_tag_excel if is_excel_dynamic and issue_tag_excel else "Vol. 2026 趨勢動態探索"
        newsletter_theme = synthesize_weekly_express_theme(articles)
        focus_domains = "全網熱點、算力基礎設施、Agentic AI 攻防、智慧製造"
        goal_desc = "【無預設主題模式】自動分析當週全網熱度最高的前 4 大真實趨勢與重磅事件。"
        print(f"🌐 [Stage 2 Curator Engine - Dynamic Discovery Mode] Automatically clustered weekly top trends!")
        print(f"   期數: {issue_tag}\n   動態推導主題: {newsletter_theme}\n   重點領域: {focus_domains}")
    else:
        newsletter_theme, issue_tag, focus_domains, goal_desc = theme_from_excel, issue_tag_excel, focus_domains_excel, goal_desc_excel
        print(f"🎯 [Stage 2 Curator Engine - Preset Theme Mode] Evaluating Stage 1 database for Active Theme: [{issue_tag}] {newsletter_theme}...")
    
    target_domains, domain_keyword_map = build_active_theme_matcher(newsletter_theme, focus_domains, goal_desc)
    print(f"🔍 Active Theme Domain Matchers: {target_domains}")
        
    curated_candidates = []
    for art in articles:
        total_score, matched_tags = evaluate_article(
            art, newsletter_theme, focus_domains, target_domains, domain_keyword_map, goal_desc
        )
            
        if total_score >= 35:
            curated_candidates.append({
                "score": total_score,
                "title": art["title"],
                "pub_date": art["pub_date"],
                "link": art["link"],
                "source": art["source"],
                "functional_tags": " / ".join(matched_tags) if matched_tags else "當週重磅新聞 / 科技趨勢",
                "description": art.get("description", art.get("summary", ""))
            })

    # Sort Tier 1 candidates by score descending
    curated_candidates.sort(key=lambda x: (x["score"], x["pub_date"]), reverse=True)
    candidate_limit = 25
    tier1_top_candidates = curated_candidates[:candidate_limit]
    print(f"🎯 [Tier 1 Candidate Filter] Filtered down to top {len(tier1_top_candidates)} candidates for Tier 2 LLM curation.")

    # Tier 2: LLM Deep Semantic Evaluation & Rationale Generation
    final_candidates = evaluate_candidates_via_llm(tier1_top_candidates, newsletter_theme, focus_domains, goal_desc)
    
    # Sort strictly by final score descending
    final_candidates.sort(key=lambda x: (x["score"], x["pub_date"]), reverse=True)
    
    # Step 3: Qualified Score Threshold (最低合格門檻，預設 80 分)
    QUALIFIED_SCORE_THRESHOLD = int(os.environ.get("QUALIFIED_SCORE_THRESHOLD", "80"))
    qualified_articles = [art for art in final_candidates if int(art["score"]) >= QUALIFIED_SCORE_THRESHOLD]
    
    # Step 4: Top N (最多取 Top 6，零湊數機制)
    MAX_TOP_COUNT = 6
    selected_articles = qualified_articles[:MAX_TOP_COUNT]

    print(f"🎯 [Stage 2 Evaluation] Total Evaluated: {len(final_candidates)} 篇 | Qualified (Score >= {QUALIFIED_SCORE_THRESHOLD}): {len(qualified_articles)} 篇 | Selected (Top {MAX_TOP_COUNT}): {len(selected_articles)} 篇")
    if len(selected_articles) == 0:
        print(f"⚠️ [Stage 2 Notice] 本期符合主題之高品質新聞不足 (無文章達到 {QUALIFIED_SCORE_THRESHOLD} 分門檻，堅持零湊數輸出 0 篇)。")
    else:
        print(f"✅ Selected top {len(selected_articles)} curated articles matching mode [{pipeline_mode}: {newsletter_theme}] (ranked by score >= {QUALIFIED_SCORE_THRESHOLD}).")

    try:
        from db_manager import save_curated_articles
        c_count = save_curated_articles(issue_tag, selected_articles)
        print(f"✨ [Stage 2 DB] Saved {c_count} curated articles for [{issue_tag}] to DB.")
    except Exception as e:
        print(f"⚠️ [Notice] Could not save curated articles to DB: {e}")

    # Code Review Item 6: Build Unified DTO for Exporters
    export_dtos = prepare_curated_export_dto(selected_articles, QUALIFIED_SCORE_THRESHOLD)

    # Export Excel workbook (data/stage2_curated_news.xlsx)
    output_xlsx = "data/stage2_curated_news.xlsx"
    wb = openpyxl.Workbook()
    
    # Sheet 1: 評比標準與原則
    ws_criteria = wb.active
    ws_criteria.title = "Stage 2 選文原則與評比架構"
    ws_criteria.views.sheetView[0].showGridLines = True
    
    h_font = Font(name="Microsoft JhengHei", size=11, bold=True, color="FFFFFF")
    h_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    b_font = Font(name="Microsoft JhengHei", size=10)
    title_head = Font(name="Microsoft JhengHei", size=14, bold=True, color="0F766E")
    
    ws_criteria.cell(row=1, column=1, value="Stage 2 本週電子報選文評比指標與架構說明").font = title_head
    ws_criteria.cell(row=2, column=1, value=f"本週電子報主題 ({issue_tag})：{newsletter_theme}").font = Font(name="Microsoft JhengHei", size=11, bold=True, color="1E3A8A")
    ws_criteria.cell(row=3, column=1, value=f"重點對應領域：{focus_domains} | 合格門檻：Score >= {QUALIFIED_SCORE_THRESHOLD} 分 (最多取 Top {MAX_TOP_COUNT}，零湊數)").font = b_font
    ws_criteria.cell(row=4, column=1, value="")
    
    eval_criteria = [
        {"name": "1. 主題契合度 (Theme Relevance)", "weight": "55%", "description": f"文章是否直接關聯本周電子報主題《{newsletter_theme}》及重點對應領域（{focus_domains}）。"},
        {"name": "2. 具體應用場景 (Specific Use Case)", "weight": "15%", "description": "是否包含明確產業落地應用場景、工作流自動化或具體營運流程。"},
        {"name": "3. 技術與架構細節 (Technical Depth)", "weight": "15%", "description": "是否涵蓋模型、演算法、雲地架構、API 或軟硬體系統細節。"},
        {"name": "4. 量化效益與 ROI 數據 (Quantified Impact & ROI)", "weight": "10%", "description": "是否提供具體數據指標（%、倍數、成本降低、良率提升等）或實務解決方案框架。"},
        {"name": "5. 媒體權威度與時效性 (Authority & Timeliness)", "weight": "5%", "description": "報導來源是否為國內外權威科技/商業媒體，且具備明確發布日期。"},
        {"name": "6. 最低合格門檻與零湊數原則 (Qualified Threshold & Zero Quota-Filling)", "weight": "門檻", "description": f"總分需達 {QUALIFIED_SCORE_THRESHOLD} 分以上方屬 Qualified；依分數排序最多取 Top {MAX_TOP_COUNT} 篇。若合格不足 6 篇或為 0 篇，絕不降低門檻湊數。"}
    ]
    
    ws_criteria.append(["評比指標項目", "指標權重", "評比標準與審核原則詳細說明"])
    for col in range(1, 4):
        c = ws_criteria.cell(row=5, column=col)
        c.font = h_font
        c.fill = h_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws_criteria.row_dimensions[5].height = 26
        
    for item in eval_criteria:
        ws_criteria.append([item["name"], item["weight"], item["description"]])
        
    for r in range(6, 6 + len(eval_criteria)):
        ws_criteria.cell(row=r, column=1).font = Font(name="Microsoft JhengHei", size=10, bold=True)
        ws_criteria.cell(row=r, column=2).font = Font(name="Consolas", size=10, bold=True, color="0F766E")
        ws_criteria.cell(row=r, column=3).font = b_font
        ws_criteria.row_dimensions[r].height = 24
        
    ws_criteria.column_dimensions['A'].width = 35
    ws_criteria.column_dimensions['B'].width = 15
    ws_criteria.column_dimensions['C'].width = 85

    # Sheet 2: 電子報精選新聞清單
    ws_news = wb.create_sheet(title="電子報精選新聞列表")
    ws_news.views.sheetView[0].showGridLines = True
    
    headers = [
        "項次", "評比得分", "職能對應標籤", "新聞標題", "新聞發布日期", 
        "正確原文連結", "媒體來源", "選入理由與電子報推薦切入點", "新聞摘要"
    ]
    ws_news.append(headers)
    
    header_fill_b = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    for col_num in range(1, len(headers) + 1):
        cell = ws_news.cell(row=1, column=col_num)
        cell.font = h_font
        cell.fill = header_fill_b
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws_news.row_dimensions[1].height = 28
    
    score_font = Font(name="Consolas", size=11, bold=True, color="D97706")
    link_font = Font(name="Microsoft JhengHei", size=10, color="0066CC", underline="single")
    
    for dto in export_dtos:
        row_idx = dto["rank"] + 1
        ws_news.append([
            dto["rank"],
            dto["score_str"],
            dto["functional_tags"],
            dto["title"],
            dto["pub_date"],
            dto["link"],
            dto["source"],
            dto["rationale"],
            dto["description"]
        ])
        
        c_num = ws_news.cell(row=row_idx, column=1)
        c_score = ws_news.cell(row=row_idx, column=2)
        c_domain = ws_news.cell(row=row_idx, column=3)
        c_title = ws_news.cell(row=row_idx, column=4)
        c_date = ws_news.cell(row=row_idx, column=5)
        c_link = ws_news.cell(row=row_idx, column=6)
        c_src = ws_news.cell(row=row_idx, column=7)
        c_rat = ws_news.cell(row=row_idx, column=8)
        c_desc = ws_news.cell(row=row_idx, column=9)
        
        c_num.alignment = Alignment(horizontal="center", vertical="center")
        c_score.font = score_font
        c_score.alignment = Alignment(horizontal="center", vertical="center")
        c_domain.font = Font(name="Microsoft JhengHei", size=10, bold=True)
        c_title.font = Font(name="Microsoft JhengHei", size=10, bold=True)
        c_date.font = Font(name="Consolas", size=10)
        c_date.alignment = Alignment(horizontal="center", vertical="center")
        c_link.font = link_font
        if dto["link"] != "-":
            c_link.hyperlink = dto["link"]
        c_src.font = b_font
        c_rat.font = b_font
        c_desc.font = b_font
        
        for col in range(1, 10):
            ws_news.cell(row=row_idx, column=col).border = thin_border
        ws_news.row_dimensions[row_idx].height = 26
        
    col_widths_b = {1: 8, 2: 12, 3: 32, 4: 45, 5: 18, 6: 48, 7: 24, 8: 55, 9: 45}
    for col_idx, width in col_widths_b.items():
        col_letter = get_column_letter(col_idx)
        ws_news.column_dimensions[col_letter].width = width

    try:
        wb.save(output_xlsx)
        print(f"✨ [Stage 2 SUCCESS] Created Excel report at: {output_xlsx}")
    except PermissionError:
        fallback_xlsx = "data/stage2_curated_news_latest.xlsx"
        try:
            wb.save(fallback_xlsx)
            print(f"⚠️ [Stage 2 Notice] {output_xlsx} 目前被其他程式（如 Excel）開啟鎖定中，已另存備份至: {fallback_xlsx}")
        except Exception:
            print(f"⚠️ [Stage 2 Notice] {output_xlsx} 目前被 Excel 開啟鎖定中，跳過覆寫，繼續產出 Markdown 報告。")

    # Export Markdown report (data/stage2_curated_report.md)
    md_path = "data/stage2_curated_report.md"
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(f"# 🎯 Stage 2 - 本週電子報精選新聞報告\n\n")
        f.write(f"> **本週電子報主題 ({issue_tag})**: `{newsletter_theme}`\n")
        f.write(f"> **重點對應領域**: `{focus_domains}`\n")
        f.write(f"> **資料來源庫**: Stage 1 累積新聞庫 (`{len(articles)}` 篇) | **合格篇數**: `{len(qualified_articles)}` 篇 | **最終選入篇數**: `{len(selected_articles)}` 篇 (門檻 Score >= {QUALIFIED_SCORE_THRESHOLD})\n\n")
        
        f.write("## 🎯 Stage 2 評比標準與權重配比\n\n")
        f.write("| 評比指標項目 | 權重 | 評比標準與原則說明 |\n| --- | --- | --- |\n")
        for item in eval_criteria:
            f.write(f"| **{item['name']}** | `{item['weight']}` | {item['description']} |\n")
        f.write("\n")
        
        f.write("## 📰 精選新聞清單與推薦切入點 (按分數高低排序，最多 Top 6，零湊數)\n\n")
        if not selected_articles:
            f.write(f"> ⚠️ **本期符合主題之高品質新聞不足**：本期所有候選報導經 Hard Constraint 與 5 大維度評比後，無任何文章達到最低合格門檻 (Score >= {QUALIFIED_SCORE_THRESHOLD} 分)。系統堅持「零湊數原則」，不降低分數門檻亦不引入低相關新聞。\n\n")
        else:
            for dto in export_dtos:
                f.write(f"### {dto['rank']}. [{dto['title']}]({dto['link']})\n")
                f.write(f"- **評比得分**: `{dto['score_str']}` | **職能對應**: `{dto['functional_tags']}`\n")
                f.write(f"- **新聞發布日期**: `{dto['pub_date']}` | **媒體來源**: `{dto['source']}`\n")
                f.write(f"- 💡 **選入理由與建議切入點**:\n  >{dto['rationale'].replace(chr(10), chr(10) + '  >')}\n")
                if dto['description']:
                    f.write(f"- **摘要**: {dto['description']}\n")
                f.write(f"- 🔗 **正確原文連結**: [{dto['link']}]({dto['link']})\n\n")
            
    print(f"📝 [Stage 2 SUCCESS] Created Markdown report at: {md_path}")

if __name__ == "__main__":
    main()
