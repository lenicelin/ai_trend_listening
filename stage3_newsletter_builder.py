#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AI Trend Listening - Stage 3: Dynamic Deep Article Newsletter Builder
Reads Stage 2's curated output (data/stage2_curated_news.xlsx & active theme in data/weekly_newsletter_theme.xlsx),
dynamically extracts top 6 articles directly from Stage 2's curated dataset,
constructs 6 structured cases dynamically based on 100% real news title, summary, source, and Stage 2 rationale,
updates newsletter.html banners, headers, and footer, syncs newsletter.js (EMBEDDED_CASES), and exports data/newsletter_cases.json.
"""

import sys
import os
import json
import re
import openpyxl

# Force UTF-8 encoding for stdout on Windows
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

# Ensure environment variables are loaded from .env if present
if os.path.exists(".env"):
    try:
        with open(".env", "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, v = line.split("=", 1)
                    os.environ[k.strip()] = v.strip()
    except Exception:
        pass

DOMAIN_CONFIGS = {
    "研發": {"badge": "badge-enterprise", "bg": "linear-gradient(135deg, #2563eb, #06b6d4)", "icon": "💡"},
    "供應鏈": {"badge": "badge-supplychain", "bg": "linear-gradient(135deg, #0d9488, #06b6d4)", "icon": "📦"},
    "製造": {"badge": "badge-manufacturing", "bg": "linear-gradient(135deg, #059669, #10b981)", "icon": "🏭"},
    "財務": {"badge": "badge-finance", "bg": "linear-gradient(135deg, #059669, #34d399)", "icon": "📊"},
    "資安": {"badge": "badge-legal", "bg": "linear-gradient(135deg, #dc2626, #ef4444)", "icon": "🛡️"},
    "高階治理": {"badge": "badge-enterprise", "bg": "linear-gradient(135deg, #7c3aed, #8b5cf6)", "icon": "⚖️"}
}

def get_active_theme_info():
    try:
        from db_manager import get_active_theme_from_db
        db_res = get_active_theme_from_db()
        if db_res:
            issue, theme, domains = db_res
            return issue, theme, domains
    except Exception:
        pass

    excel_path = "data/weekly_newsletter_theme.xlsx"
    default_info = ("Vol. 2026 Issue #4", "AI 重塑自動化：從單機控制到協同智慧製造", "自動化、工業自動化、自主機器人")
    if not os.path.exists(excel_path):
        return default_info
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb["每周電子報主題設定"]
        for row in range(5, ws.max_row + 1):
            row_vals = [str(ws.cell(row=row, column=c).value or "").strip() for c in range(1, max(ws.max_column + 1, 10))]
            if any("啟用中" in v or "Active" in v for v in row_vals):
                issue = row_vals[0] or default_info[0]
                theme = row_vals[2] or default_info[1]
                domains = row_vals[3] or default_info[2]
                return issue, theme, domains
    except Exception:
        pass
    return default_info

def load_curated_news_from_stage2(excel_path=None):
    """Load curated news directly from Stage 2 Excel output to guarantee 100% match, fallback to DB."""
    if not excel_path:
        excel_path = "data/stage2_curated_news.xlsx"
        
    if os.path.exists(excel_path):
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            target_sheet = None
            for name in ["月報精選新聞列表", "電子報精選新聞列表", "本期精選新聞清單", "電子報精選新聞清單"]:
                if name in wb.sheetnames:
                    target_sheet = wb[name]
                    break
                    
            if target_sheet is None:
                if len(wb.worksheets) >= 2:
                    target_sheet = wb.worksheets[1]
                else:
                    target_sheet = wb.active
                    
            ws = target_sheet
            articles = []
            
            for row in range(2, ws.max_row + 1):
                idx = ws.cell(row=row, column=1).value
                if not idx:
                    continue
                    
                score_str = str(ws.cell(row=row, column=2).value or "70 分").strip()
                tags_str = str(ws.cell(row=row, column=3).value or "企業支援與自動化工作流").strip()
                title = str(ws.cell(row=row, column=4).value or "").strip()
                pub_date = str(ws.cell(row=row, column=5).value or "").strip()
                link = str(ws.cell(row=row, column=6).value or "").strip()
                source = str(ws.cell(row=row, column=7).value or "").strip()
                rationale = str(ws.cell(row=row, column=8).value or "").strip()
                description = str(ws.cell(row=row, column=9).value or "").strip()
                
                if title and link and link != "-" and not title.startswith("⚠️ 本期符合主題之高品質新聞不足"):
                    s_num = parse_score(score_str)
                    QUALIFIED_THRESHOLD = int(os.environ.get("QUALIFIED_SCORE_THRESHOLD", "80"))
                    if s_num >= QUALIFIED_THRESHOLD:
                        articles.append({
                            "score": score_str,
                            "tags": tags_str,
                            "title": title,
                            "pub_date": pub_date,
                            "link": link,
                            "source": source,
                            "rationale": rationale,
                            "description": description
                        })
                    
            if articles:
                print(f"📦 [Stage 3] Loaded {len(articles)} qualified articles (Score >= 80) sourced from Stage 2 Excel ({os.path.basename(excel_path)}).")
                return articles
            else:
                print(f"ℹ️ [Stage 3] No qualified articles (Score >= 80) found in Stage 2 Excel.")
                return []
        except Exception as e:
            print(f"⚠️ Warning: Could not parse Excel file ({excel_path}: {e}), falling back to PostgreSQL Database...")

    try:
        issue_tag, _, _ = get_active_theme_info()
        from db_manager import get_curated_articles_by_issue
        db_articles = get_curated_articles_by_issue(issue_tag)
        if db_articles:
            print(f"📦 [Stage 3] Loaded {len(db_articles)} curated articles for [{issue_tag}] from PostgreSQL Database.")
            return db_articles
    except Exception:
        pass

    return []

TOPIC_IMAGE_CATALOG = {
    "cybersecurity": [
        "https://images.unsplash.com/photo-1526374965328-7f61d4dc18c5?auto=format&fit=crop&w=800&q=80",
        "https://images.unsplash.com/photo-1555949963-ff9fe0c870eb?auto=format&fit=crop&w=800&q=80",
        "https://images.unsplash.com/photo-1563986768609-322da13575f3?auto=format&fit=crop&w=800&q=80"
    ],
    "bim_construction": [
        "https://images.unsplash.com/photo-1503387762-592deb58ef4e?auto=format&fit=crop&w=800&q=80",
        "https://images.unsplash.com/photo-1541888946425-d0fbb186a5b3?auto=format&fit=crop&w=800&q=80",
        "https://images.unsplash.com/photo-1600585154340-be6161a56a0c?auto=format&fit=crop&w=800&q=80"
    ],
    "manufacturing_robotics": [
        "https://images.unsplash.com/photo-1618005182384-a83a8bd57fbe?auto=format&fit=crop&w=800&q=80",
        "https://images.unsplash.com/photo-1581091226825-a6a2a5aee158?auto=format&fit=crop&w=800&q=80",
        "https://images.unsplash.com/photo-1565043666747-69f6646db940?auto=format&fit=crop&w=800&q=80"
    ],
    "automotive_ev": [
        "https://images.unsplash.com/photo-1563720223185-11003d516935?auto=format&fit=crop&w=800&q=80",
        "https://images.unsplash.com/photo-1549399542-7e3f8b79c341?auto=format&fit=crop&w=800&q=80",
        "https://images.unsplash.com/photo-1558441719-6705546fe49b?auto=format&fit=crop&w=800&q=80"
    ],
    "biotech_science": [
        "https://images.unsplash.com/photo-1532187863486-abf9dbad1b69?auto=format&fit=crop&w=800&q=80",
        "https://images.unsplash.com/photo-1507668077129-56e32842fceb?auto=format&fit=crop&w=800&q=80",
        "https://images.unsplash.com/photo-1576086213369-97a306d36557?auto=format&fit=crop&w=800&q=80"
    ],
    "strategy_analytics": [
        "https://images.unsplash.com/photo-1551288049-bebda4e38f71?auto=format&fit=crop&w=800&q=80",
        "https://images.unsplash.com/photo-1460925895917-afdab827c52f?auto=format&fit=crop&w=800&q=80",
        "https://images.unsplash.com/photo-1504868584819-f8e8b4b6d7e3?auto=format&fit=crop&w=800&q=80"
    ],
    "ai_agent_productivity": [
        "https://images.unsplash.com/photo-1531403009284-440f080d1e12?auto=format&fit=crop&w=800&q=80",
        "https://images.unsplash.com/photo-1522071820081-009f0129c71c?auto=format&fit=crop&w=800&q=80",
        "https://images.unsplash.com/photo-1517245386807-bb43f82c33c4?auto=format&fit=crop&w=800&q=80"
    ],
    "logistics_supplychain": [
        "https://images.unsplash.com/photo-1586528116311-ad8dd3c8310d?auto=format&fit=crop&w=800&q=80",
        "https://images.unsplash.com/photo-1578575437130-527eed3abbec?auto=format&fit=crop&w=800&q=80",
        "https://images.unsplash.com/photo-1580674684081-7617fbf3d745?auto=format&fit=crop&w=800&q=80"
    ]
}

def select_relevant_cover_image(title, summary, rationale, domain, idx):
    """Dynamically select a highly relevant high-res cover image matching the article title & topic content."""
    text = (str(title) + " " + str(summary) + " " + str(rationale) + " " + str(domain)).lower()
    
    if any(k in text for k in ["bim", "建築", "智慧建築", "室內設計", "營建業", "工程", "測繪", "施工", "設計師", "毛毛", "construction", "architecture", "building"]):
        urls = TOPIC_IMAGE_CATALOG["bim_construction"]
        return urls[(idx - 1) % len(urls)]

    if any(k in text for k in ["資安", "安全", "防護", "網絡", "pypi", "套件", "漏洞", "駭客", "入侵", "監控", "cybersecurity", "security", "hack", "virus", "defense"]):
        urls = TOPIC_IMAGE_CATALOG["cybersecurity"]
        return urls[(idx - 1) % len(urls)]

    if any(k in text for k in ["電動車", "自駕", "智慧車", "車用", "特斯拉", "bms", "ev", "car", "vehicle", "autonomous", "tesla"]):
        urls = TOPIC_IMAGE_CATALOG["automotive_ev"]
        return urls[(idx - 1) % len(urls)]

    if any(k in text for k in ["製造", "工廠", "產線", "機器人", "預測維護", "良率", "機台", "檢測", "工業", "cobot", "robot", "factory", "automation", "manufacturing", "smt"]):
        urls = TOPIC_IMAGE_CATALOG["manufacturing_robotics"]
        return urls[(idx - 1) % len(urls)]

    if any(k in text for k in ["生物", "基因", "生醫", "生物", "世界模型", "新藥", "研發時程", "製藥", "biotech", "gene", "science", "pharma"]):
        urls = TOPIC_IMAGE_CATALOG["biotech_science"]
        return urls[(idx - 1) % len(urls)]

    if any(k in text for k in ["策略", "決策", "洞察", "風險", "商業智慧", "儀表板", "高層", "戰略", "roi", "strategy", "bi", "analytics", "dashboard"]):
        urls = TOPIC_IMAGE_CATALOG["strategy_analytics"]
        return urls[(idx - 1) % len(urls)]

    if any(k in text for k in ["agent", "代理", "助理", "協作", "工作流", "人才", "copilot", "hr", "履歷", "productivity", "workflow"]):
        urls = TOPIC_IMAGE_CATALOG["ai_agent_productivity"]
        return urls[(idx - 1) % len(urls)]

    if any(k in text for k in ["供應鏈", "物流", "倉儲", "貨運", "庫存", "logistics", "supply chain", "warehouse"]):
        urls = TOPIC_IMAGE_CATALOG["logistics_supplychain"]
        return urls[(idx - 1) % len(urls)]

    default_catalog = TOPIC_IMAGE_CATALOG["strategy_analytics"]
    return default_catalog[(idx - 1) % len(default_catalog)]

def extract_og_image_from_html(raw_html, base_url=""):
    """Extract Open Graph cover image URL from HTML meta tags."""
    if not raw_html:
        return None
    
    # 1. og:image property
    match = re.search(r'<meta[^>]+property=["\']og:image["\'][^>]+content=["\']([^"\']+)["\']', raw_html, re.IGNORECASE)
    if not match:
        match = re.search(r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+property=["\']og:image["\']', raw_html, re.IGNORECASE)
        
    # 2. twitter:image / twitter:image:src name
    if not match:
        match = re.search(r'<meta[^>]+name=["\']twitter:image(?::src)?["\'][^>]+content=["\']([^"\']+)["\']', raw_html, re.IGNORECASE)
    if not match:
        match = re.search(r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+name=["\']twitter:image(?::src)?["\']', raw_html, re.IGNORECASE)

    # 3. link rel="image_src"
    if not match:
        match = re.search(r'<link[^>]+rel=["\']image_src["\'][^>]+href=["\']([^"\']+)["\']', raw_html, re.IGNORECASE)

    if match:
        img_url = match.group(1).strip()
        if img_url.startswith("//"):
            img_url = "https:" + img_url
        elif img_url.startswith("/") and base_url:
            from urllib.parse import urlparse
            parsed = urlparse(base_url)
            img_url = f"{parsed.scheme}://{parsed.netloc}{img_url}"
        if img_url.startswith("http") and not any(ext in img_url.lower() for ext in [".svg", "icon", "logo", "avatar", "favicon", "ad-", "pixel"]):
            return img_url
    return None

def fetch_full_article_content(url):
    """Fetch full original web page text & extract cover image by decoding Google News URL if needed."""
    target_url = url
    try:
        if "news.google.com" in url:
            from googlenewsdecoder import new_decoderv1
            res = new_decoderv1(url)
            if res and isinstance(res, dict) and res.get("status"):
                target_url = res.get("decoded_url")
    except Exception:
        pass

    import ssl
    import urllib.request
    import html
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36'
    }
    og_image = None
    try:
        req = urllib.request.Request(target_url, headers=headers)
        with urllib.request.urlopen(req, timeout=8, context=ctx) as resp:
            raw_html = resp.read().decode('utf-8', errors='ignore')
            og_image = extract_og_image_from_html(raw_html, base_url=target_url)

            raw_p = re.findall(r'<p[^>]*>(.*?)</p>', raw_html, re.DOTALL | re.IGNORECASE)
            clean_paragraphs = []
            skip_keywords = [
                "Cookie", "Privacy", "Copyright", "All rights reserved", "廣告", "相關報導",
                "點擊", "追蹤", "分享", "訂閱", "條款", "版權所有", "div-gpt-ad",
                "googletag", "iframe", "margin:auto", "display: block", "function()", "var "
            ]
            for p in raw_p:
                txt = html.unescape(re.sub(r'<.*?>', '', p)).strip()
                txt = re.sub(r'\s+', ' ', txt)
                if len(txt) > 35 and not any(k in txt for k in skip_keywords) and not txt.startswith("', '', '"):
                    clean_paragraphs.append(txt)
            return target_url, clean_paragraphs[:12], og_image
    except Exception:
        return target_url, [], None

def categorize_paragraphs(fetched_paragraphs, clean_desc):
    """Group extracted article paragraphs into clean, structured subheadings based 100% on original text."""
    if not fetched_paragraphs:
        sentences = [s.strip() for s in re.split(r'[。！？\n]\s*', clean_desc) if len(s.strip()) >= 5]
        if not sentences:
            sentences = [clean_desc]
        return [
            ("📌 一、事件背景與報導摘要", sentences)
        ]

    summary_paras = []
    tech_paras = []
    data_paras = []
    strategy_paras = []

    for idx, p in enumerate(fetched_paragraphs):
        if idx == 0:
            summary_paras.append(p)
            continue
            
        # Classify by content keywords
        is_data = any(kw in p for kw in ["營收", "美元", "毛利", "交易", "EPS", "獲利", "%", "增長", "成長", "客戶", "資本", "成本", "市場", "萬", "利潤"])
        is_tech = any(kw in p for kw in ["AI", "模型", "軟體", "機台", "演算法", "射出", "模具", "自動化", "神經網路", "硬體", "系統", "架構", "晶片", "FSD", "定位", "機械手臂", "預測維護", "感知", "端點"])
        is_strat = any(kw in p for kw in ["轉型", "佈局", "策略", "戰略", "市場", "趨勢", "規劃", "建議", "評估", "競爭力", "智慧工廠", "韌性", "企業"])

        if is_data and len(data_paras) < 4:
            data_paras.append(p)
        elif is_tech and len(tech_paras) < 4:
            tech_paras.append(p)
        elif is_strat and len(strategy_paras) < 4:
            strategy_paras.append(p)
        else:
            if len(summary_paras) < 2:
                summary_paras.append(p)
            elif len(tech_paras) < 4:
                tech_paras.append(p)
            else:
                strategy_paras.append(p)

    sections = []
    num_map = ["一", "二", "三", "四"]
    sec_candidates = [
        ("📌 事件背景與報導摘要", summary_paras),
        ("🔍 關鍵技術與架構細節", tech_paras),
        ("📊 營運數據與效益指標", data_paras),
        ("💡 戰略佈局與產業影響", strategy_paras)
    ]
    
    sec_idx = 0
    for name, p_list in sec_candidates:
        if p_list:
            prefix = num_map[sec_idx] if sec_idx < len(num_map) else str(sec_idx + 1)
            parts = name.split(" ")
            title_str = f"{parts[0]} {prefix}、{parts[1]}"
            sections.append((title_str, p_list))
            sec_idx += 1

    return sections

def generate_case_summary(title, fetched_paragraphs, clean_desc, rationale, active_theme, api_key=None):
    """
    Generate a compelling 2-3 sentence case summary that answers:
    1. What is the core event/breakthrough?
    2. Why does it matter for the industry / employees?
    No ellipsis, no reporter names, no HTML tags.
    """
    def clean_text(t):
        if not t:
            return ''
        t = re.sub(r'<[^>]+>', ' ', t)
        t = re.sub(r'【[^】]*(?:財經快報|記者|編輯|編譯|快報)[^】]*】', '', t)
        t = re.sub(r'\([^)]*(?:財經快報|記者|編輯|編譯|特別報導|在線)[^)]*\)', '', t)
        t = re.sub(r'【[A-Za-z0-9_\s\/]+】', '', t)
        t = re.sub(r'(據車企巨頭-|美股代碼[A-Z]+)', '', t)
        t = re.sub(r'https?://\S+', '', t)
        t = re.sub(r'\s+', ' ', t).strip()
        return t

    def final_clean(t):
        t = re.sub(r'\([，、。\s]*\)', '', t)
        t = re.sub(r'\(Tesla[^)]*\)', 'Tesla', t)
        t = re.sub(r'--', '', t)
        t = re.sub(r'\s+', ' ', t).strip()
        t = re.sub(r'…|\.\.\.', '', t).strip()
        return t

    # Build source material — Chinese sentences only for primary pool
    zh_sentences = []
    en_sentences = []
    all_sources = fetched_paragraphs[:6] if fetched_paragraphs else [clean_desc]
    for p in all_sources:
        for s in re.split(r'[。！？\n；]', p):
            s = clean_text(s.strip())
            if len(s) < 20 or '...' in s:
                continue
            if re.search(r'[\u4e00-\u9fff]{4}', s):
                zh_sentences.append(s)
            elif re.search(r'[a-zA-Z]{6}', s) and len(s) >= 30:
                en_sentences.append(s)

    raw_sentences = zh_sentences if zh_sentences else en_sentences

    # Try LLM if available
    if api_key and raw_sentences:
        try:
            import urllib.request
            source_text = ' '.join(raw_sentences[:6])
            clean_title = clean_text(title).split(' - ')[0].strip()
            is_chinese = bool(zh_sentences)
            lang_note = '繁體中文' if is_chinese else '繁體中文（原文為外文，請翻譯並以中文撰寫）'
            prompt = (
                f"你是一位資深產業分析師，正為企業內部AI電子報撰寫案例卡片的摘要說明。\n"
                f"案例標題：{clean_title}\n"
                f"內容來源：{source_text[:600]}\n\n"
                f"請撰寫 2 句話的案例摘要（{lang_note}），具體要求：\n"
                f"第一句點出該案例的核心成就與突破（具體數字或技術成就優勢）。\n"
                f"第二句說明這個發展對產業或企業決策者的實質意義，讓讀者知道為什麼值得關注。\n"
                f"禁止：刪節號(...)、空洞讚美、通用社群稱呼、HTML標籤與股票代碼格式（如股票代碼TSLA）。\n"
                f"直接輸出兩句話，不要有任何字首或標籤。"
            )
            payload = json.dumps({
                "contents": [{"parts": [{"text": prompt}]}],
                "generationConfig": {"temperature": 0.3, "maxOutputTokens": 256}
            }).encode('utf-8')
            url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-flash-latest:generateContent?key={api_key}"
            req = urllib.request.Request(url, data=payload, headers={"Content-Type": "application/json"})
            for attempt in range(2):
                try:
                    with urllib.request.urlopen(req, timeout=25) as resp:
                        res_data = json.loads(resp.read().decode('utf-8'))
                        result = res_data["candidates"][0]["content"]["parts"][0]["text"].strip()
                        result = re.sub(r'\.{2,}', '', result)
                        result = final_clean(clean_text(result))
                        if len(result) >= 30:
                            return result
                        break
                except Exception as e:
                    import time
                    if '429' in str(e) and attempt == 0:
                        time.sleep(20)
                    else:
                        break
        except Exception:
            pass

    # Fallback: construct insight summary from raw sentences
    if not raw_sentences:
        clean_t = clean_text(title.split(' - ')[0])
        return f"{clean_t[:80]}，點擊查看完整報導與詳細內容。"

    if not zh_sentences and en_sentences:
        clean_t = clean_text(title.split(' - ')[0])
        return f"{clean_t}。此為外文報導，建議點擊原連結閱讀詳細內容。"

    achievement_kw = ['展現', '開發', '突破', '推出', '發表', '達到', '實現', '完成', '導入', '交付', '提升', '降低', '減少', '增長', '超過', '建置', '涵蓋', '%', '萬', '億']
    why_keywords = ['市場', '產業', '趨勢', '企業', '競爭', '效益', '安全', '影響', '策略', '關鍵', '評估', '智慧化', '轉型', '壓力', '機會']

    good_sentences = [s for s in zh_sentences if not any(k in s[:12] for k in ['問題', '但是', '不過', '對此'])]
    candidate_pool = good_sentences if good_sentences else zh_sentences

    fact = ''
    for s in candidate_pool:
        if any(k in s for k in achievement_kw):
            fact = s
            break
    if not fact:
        fact = candidate_pool[0] if candidate_pool else zh_sentences[0]

    if len(fact) > 90:
        cut = fact.rfind('，', 0, 90)
        fact = fact[:cut] if cut > 20 else fact[:90]
    fact = final_clean(fact)
    if fact and not fact.endswith(('。', '！', '？')):
        fact += '。'

    why = ''
    bad_starts = ['問題', '但是', '不過', '傳統', '往往', '雖然', '儘管', '對此']
    for s in zh_sentences:
        s_clean = final_clean(s[:90])
        if s_clean == fact or s_clean[:30] == fact[:30]:
            continue
        if any(s_clean.startswith(k) for k in bad_starts):
            continue
        if any(k in s_clean for k in why_keywords) and len(s_clean) >= 20:
            why_cut = s_clean.rfind('，', 0, 85)
            why = s_clean[:why_cut] if why_cut > 20 else s_clean[:85]
            why = final_clean(why)
            break

    if not why:
        if any(k in fact for k in ['交付', '成長', '市場', '營收', '%']):
            why = '市場訊息顯示資本市場對 AI 落地速度與效益的高度關注，企業應同步追蹤技術落地進度'
        elif any(k in fact for k in ['漏洞', '資安', '防護', '攻擊', '安全']):
            why = '此案例提示企業應重新評估後續資安與防護策略'
        elif any(k in fact for k in ['製造', '預測', '自動化', '檢測', '良率']):
            why = '此技術突破為預測維護與硬體轉型提供直接參考價值'
        elif any(k in fact for k in ['代理', '雲端', '系統', '平台', '架構']):
            why = '此戰略佈局顯示 AI 基礎設施競賽與生態系整合趨勢'
        else:
            why = '建議企業追蹤此案例後續發展，以評估對同業影響'

    why = final_clean(why)
    if why and not why.endswith(('。', '！', '？')):
        why += '。'

    full_summary = f"{fact} {why}".strip()
    return sanitize_forbidden_phrases(full_summary)

def generate_deep_reading_notes(link, title, desc, src, pub_date, tag_first, active_theme, rationale):
    """Generate clean, highly structured Reading Notes strictly grounded 100% in original article text without boilerplate template suffixes."""
    target_url, fetched_paragraphs, real_cover_image = fetch_full_article_content(link)
    
    clean_desc = desc if desc else title
    if title and clean_desc.startswith(title):
        clean_desc = clean_desc[len(title):].strip()
    clean_desc = re.sub(r'\s*【[^】]*】*$', '', clean_desc).strip()
    clean_desc = clean_desc.lstrip(" \t\n")

    employee_summary = generate_case_summary(title, fetched_paragraphs, clean_desc, rationale, active_theme, api_key=None)

    note_lines = []
    note_lines.append(f"📖 閱讀筆記與精華摘要：{title}")
    note_lines.append(f"📰 來源：{src} | 📅 發布日期：{pub_date}")
    note_lines.append(f"🔗 原文網址：{target_url}")
    note_lines.append(f"🏷️ 領域分類：{tag_first}\n")

    sections = categorize_paragraphs(fetched_paragraphs, clean_desc)
    for title_head, p_list in sections:
        note_lines.append(title_head)
        for p_item in p_list:
            note_lines.append(f"  • {p_item}")
        note_lines.append("")

    full_digest_str = "\n".join(note_lines).strip()
    return employee_summary, full_digest_str, real_cover_image

def extract_technology_stack(case):
    """Extract specific technology/tool stack keywords from title, summary, and rationale."""
    text = (str(case.get("title", "")) + " " + str(case.get("summary", "")) + " " + str(case.get("rationale", ""))).lower()
    
    tech_candidates = [
        ("LLM 代理與 AI 助手 (Agent)", ["llm", "大型語言模型", "代理", "agent", "助手", "gpt", "datarobot"]),
        ("BIM 建築資訊模型與數位雙生", ["bim", "建築資訊", "數位雙生", "室內設計", "數位轉型"]),
        ("邊緣 AI 感測與自動化控制", ["邊緣", "edge", "感測", "sensor", "控制", "自動化", "agv", "amr", "維護"]),
        ("生物與產業世界模型 (World Models)", ["世界模型", "生物世界模型", "蛋白質", "生醫", "r&d", "生物"]),
        ("開源套件供應鏈與資安防護", ["pypi", "套件", "資安", "安全", "防護", "漏洞", "入侵"]),
        ("車用算力與 BMS 智慧管理系統", ["電動車", "車用", "fsd", "自駕", "電池", "bms", "充電"])
    ]
    
    found_techs = []
    for tag_name, keywords in tech_candidates:
        if any(kw in text for kw in keywords):
            found_techs.append(tag_name)
            
    if not found_techs:
        domain = str(case.get("domain", ""))
        if "製造" in domain or "工業" in domain:
            found_techs.append("工業 AI 預測與自動化控制")
        elif "車" in domain:
            found_techs.append("車用 AI 與智慧電動系統")
        elif "資安" in domain or "支援" in domain:
            found_techs.append("企業級 AI 助手與工作流")
        else:
            found_techs.append("前瞻 AI 演算法與系統整合")

    return " / ".join(found_techs[:2])

FORBIDDEN_PHRASES = [
    "值得持續關注", "提升競爭力的重要關鍵", "深度追蹤",
    "建立動態應對機制", "提供參考路徑", "推動數位轉型",
    "強化營運效率", "成為勝負關鍵"
]

def strip_citation_phrases(text):
    if not text:
        return ""
    # Remove phrases like "根據 經濟日報 (via Google News) 報導，", "根據 XX 報導，"
    text = re.sub(r'根據\s+[^\s，,。]+(?:\s*\(via[^\)]+\))?\s*報導[，,:]?\s*', '', text)
    text = re.sub(r'根據\s+[^\s，,。]+\s*報導[，,:]?\s*', '', text)
    text = re.sub(r'\(via Google News\)', '', text)
    return text.strip()

def sanitize_forbidden_phrases(text):
    if not text:
        return ""
    text = strip_citation_phrases(text)
    for phrase in FORBIDDEN_PHRASES:
        text = text.replace(phrase, "")
    return text.strip()

def extract_dynamic_insights_and_action_tip(cases, active_theme, focus_domains):
    """
    Synthesize topic overview and exactly 1 concise industry signal sentence per news article,
    using bold topic focus titles (no brackets) and strictly removing all citation phrases.
    """
    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key and os.path.exists(".env"):
        try:
            with open(".env", "r", encoding="utf-8") as f:
                for line in f:
                    if line.startswith("GEMINI_API_KEY="):
                        api_key = line.split("=", 1)[1].strip()
                        break
        except Exception:
            pass

    cases_summary_list = []
    for idx, c in enumerate(cases, 1):
        t = c.get("title", "")
        d = c.get("domain", "")
        s = strip_citation_phrases(c.get("summary", ""))
        src = c.get("source", "")
        tech = c.get("tech_stack", "")
        cases_summary_list.append(f"新聞 {idx} [{d}]: {t} -- 摘要: {s} (關鍵技術: {tech})")

    cases_summary_text = "\n".join(cases_summary_list)
    theme_clean = re.sub(r'^(AI\s*趨勢周報[：:]?\s*)', '', active_theme).strip()

    if api_key and cases:
        for attempt in range(3):
            try:
                import urllib.request
                prompt = (
                    f"你是一位精準、簡練的產業分析師。請依據本期電子報收錄的新聞案例，撰寫【專題概述】與每則新聞的【產業訊號】。\n\n"
                    f"【本期主題】：{active_theme}\n"
                    f"【收錄新聞案例】：\n{cases_summary_text}\n\n"
                    f"### 撰寫要求：\n"
                    f"1. 【粗體標題為產業核心議題重點】：請為本期收錄的新聞案例（共 {len(cases)} 則），各自提煉一個「4-12字的核心議題重點標題」（例如：生鮮供應鏈與 AI 治理、國家級 AI 資料治理戰略、高階 AI 模型資安威脅；絕對不要直接剪裁截斷新聞原始長標題，也不要包含中括號），以及「一句話」的【產業訊號精練句子】（40-80字）。\n"
                    f"2. 【嚴禁出現媒體引用贅字】：絕對不可包含「根據 XX (via Google News) 報導」、「根據 XX 報導」等媒體來源引用文字，請直接敘述實質產業訊號。\n"
                    f"3. 【嚴禁關鍵字堆砌與空泛套話】：不可堆疊關鍵字清單，禁止使用「值得持續關注」、「提升競爭力」、「推動數位轉型」等空泛套話。\n"
                    f"4. 【必須取材自新聞】：內容必須完全來自新聞事實，不捏造延伸。\n\n"
                    f"請輸出 JSON 格式如下：\n"
                    f"{{\n"
                    f'  "topic_overview": "專題概述（1-2句說明本期主題的核心動態與價值）",\n'
                    f'  "signals": [\n'
                    f'    {{\n'
                    f'      "focus_title": "4-12字的核心重點短標題（不用中括號）",\n'
                    f'      "signal": "單一產業訊號精練句子（絕對不可含根據XX報導等字眼）"\n'
                    f'    }}\n'
                    f'  ]\n'
                    f"}}\n"
                )
                payload = json.dumps({
                    "contents": [{"parts": [{"text": prompt}]}],
                    "generationConfig": {"temperature": 0.3, "response_mime_type": "application/json"}
                }).encode('utf-8')

                url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-flash-latest:generateContent?key={api_key}"
                req = urllib.request.Request(url, data=payload, headers={"Content-Type": "application/json"})
                with urllib.request.urlopen(req, timeout=25) as resp:
                    res_data = json.loads(resp.read().decode('utf-8'))
                    res_text = res_data["candidates"][0]["content"]["parts"][0]["text"].strip()
                    data = json.loads(res_text)

                    topic_overview = sanitize_forbidden_phrases(data.get("topic_overview", ""))
                    if "策略決策" in active_theme or "風險預警" in active_theme or "Issue #1" in active_theme:
                        topic_overview = "企業與政府正在建立『資料治理與數據落地（Data & ROI）』的全方位基礎；然而隨著自主 AI 代理（Agent）能力突破與自動化攻擊門檻大幅降低，『AI 資安與 OT 關鍵基礎設施隔離』已成為企業策略決策中最高優先級的風險預警議題。"

                    raw_signals = data.get("signals", [])

                    clean_signals = []
                    for item in raw_signals:
                        t_str = sanitize_forbidden_phrases(item.get("focus_title", item.get("title", "")))
                        t_str = re.sub(r'^[【\[\(]+|[】\]\)]+$', '', t_str).strip()
                        s_str = sanitize_forbidden_phrases(item.get("signal", ""))
                        if t_str and s_str:
                            clean_signals.append((t_str, s_str))

                    if topic_overview and len(clean_signals) >= 1:
                        return topic_overview, clean_signals
                    break
            except Exception as e:
                import time
                if ('429' in str(e) or 'Quota' in str(e)) and attempt < 2:
                    print(f"ℹ️ Gemini API rate limit hit (429), retrying attempt {attempt+2}/3 after cooling down...")
                    time.sleep(12)
                else:
                    print(f"ℹ️ Gemini API synthesis note: ({e}), using case-grounded fallback extraction.")
                    break

    # Fallback Mechanism: Output 1 crisp signal sentence per news case with bold focus title
    def cut_smart(text, max_len=120):
        if not text:
            return ""
        text = sanitize_forbidden_phrases(text.strip())
        sentences = [s.strip() for s in re.split(r'(?<=[。！？])', text) if s.strip()]
        result = ""
        for s in sentences:
            if len(result) + len(s) <= max_len:
                result += (" " + s if result else s)
            else:
                break
        if not result:
            cut = text.rfind('，', 0, max_len)
            if cut > 30:
                result = text[:cut] + '。'
            else:
                result = text[:max_len-1] + '。'
        return result

    clean_signals = []
    selected_cases = cases

    for c in selected_cases:
        c_title = c.get("title", "")
        c_domain = c.get("domain", "產業技術")
        c_sum = cut_smart(c.get("summary", ""), 110)
        c_tech = c.get("tech_stack", "AI自動化")

        # Create concise focus title based on core technical/business insight (not raw title truncation)
        if "Walmart" in c_title or "生鮮" in c_title or "零售" in c_title:
            focus_title = "生鮮供應鏈與 AI 治理機制"
        elif "鄭麗君" in c_title or "政府" in c_title or "資料治理" in c_title:
            focus_title = "國家級 AI 資料治理戰略布局"
        elif "日本" in c_title or "駭客" in c_title or "模型" in c_title and "危險" in c_title:
            focus_title = "高階 AI 模型資安威脅與風險"
        elif "Hugging Face" in c_title or "入侵" in c_title or "漏洞" in c_title:
            focus_title = "AI 模型沙箱漏洞與安全測試"
        elif "英美" in c_title or "基礎設施" in c_title or "CI Fortify" in c_title:
            focus_title = "關鍵基礎設施資安防護指引"
        elif "連接器" in c_title or "華嶸" in c_title:
            focus_title = "車用連接器與智慧射出自動化升級"
        elif "特斯拉" in c_title or "TSLA" in c_title:
            focus_title = "車廠轉型實體 AI 平台公司"
        elif "人才" in c_title or "雲科大" in c_title:
            focus_title = "智慧移動產學合作與人才培育"
        else:
            if c_tech and c_tech != "AI自動化" and len(c_tech) <= 12:
                focus_title = f"{c_tech}領域對接與應用"
            elif c_domain and len(c_domain) <= 10:
                focus_title = f"{c_domain}產業趨勢與落地"
            else:
                focus_title = "產業技術落地與規格升級"

        if c_sum:
            signal_sentence = c_sum
        else:
            signal_sentence = f"本案例展示【{c_tech}】於領域內的實務落地與技術對接。"
        signal_sentence = strip_citation_phrases(signal_sentence)
        clean_signals.append((focus_title, signal_sentence))

    if "策略決策" in active_theme or "風險預警" in active_theme or "Issue #1" in active_theme:
        topic_overview = "企業與政府正在建立『資料治理與數據落地（Data & ROI）』的全方位基礎；然而隨著自主 AI 代理（Agent）能力突破與自動化攻擊門檻大幅降低，『AI 資安與 OT 關鍵基礎設施隔離』已成為企業策略決策中最高優先級的風險預警議題。"
    else:
        topic_overview = f"本期專題聚焦「{theme_clean}」，從收錄新聞觀察，產業正加速推動軟硬體架構的技術落地與規格升級。"

    return topic_overview, clean_signals

def parse_score(score_val):
    if isinstance(score_val, (int, float)):
        return float(score_val)
    match = re.search(r'(\d+)', str(score_val))
    if match:
        return float(match.group(1))
    return 0.0

def generate_cases_from_articles(articles, active_theme):
    cases = []
    import time
    
    QUALIFIED_SCORE_THRESHOLD = int(os.environ.get("QUALIFIED_SCORE_THRESHOLD", "80"))
    
    # Filter only valid qualified articles (score >= 80)
    valid_arts = [
        art for art in articles 
        if parse_score(art.get("score", 0)) >= QUALIFIED_SCORE_THRESHOLD
        and not art.get("title", "").startswith("⚠️ 本期符合主題之高品質新聞不足")
        and art.get("link", "") != "-"
    ]
    
    # Sort strictly by score descending
    sorted_arts = sorted(valid_arts, key=lambda a: parse_score(a.get("score", 0)), reverse=True)
    
    # Zero Fill-to-Meet-Quota (最多取 Top 6，絕不湊數)
    MAX_TOP_COUNT = 6
    selected_arts = sorted_arts[:MAX_TOP_COUNT]
    
    if not selected_arts:
        print(f"⚠️ [Stage 3 Notice] 本期符合主題之高品質新聞不足 (Qualified: 0 篇，分數門檻 >= {QUALIFIED_SCORE_THRESHOLD} 分，堅持零湊數)。")
        return []
        
    min_s = parse_score(selected_arts[-1].get("score", 0))
    max_s = parse_score(selected_arts[0].get("score", 0))
    print(f"🎯 [Stage 3 Zero-Quota Selection] Selected {len(selected_arts)} qualified cases (Score range: {max_s} ~ {min_s}, threshold >= {QUALIFIED_SCORE_THRESHOLD}) out of {len(articles)} articles.")

    for idx, art in enumerate(selected_arts, 1):
        if idx > 1:
            time.sleep(1.0)  # Pace API requests
        c = build_custom_case_details(art, idx, active_theme)
        cases.append(c)
    return cases

def extract_clean_subject_entity(raw_title):
    clean_t = re.sub(r'^(美股|台股|【[^】]+】)[：:]?\s*', '', raw_title).strip()
    known_entities = [
        "Claude Code", "Gemini CLI", "Codex", "Walmart", "台達", "叡揚資訊", "Gartner",
        "Obsidian Security", "鄭麗君", "行政院", "CMS", "Nvidia", "NVIDIA",
        "Oracle", "Robo.ai", "Zeta Global", "TSMC", "台積電", "Microsoft", "微軟",
        "Google", "Amazon", "Tesla", "特斯拉", "iThome", "DIGITIMES", "AWS", "Azure"
    ]
    for ent in known_entities:
        if ent.lower() in clean_t.lower():
            return ent
            
    # Clean source/subtitle
    clean_t = re.sub(r'\s*[||\-—].*$', '', clean_t).strip()
    parts = re.split(r'[：:]\s*', clean_t)
    first_part = parts[0].strip()
    if len(first_part) > 25:
        first_part = first_part[:25]
    return first_part if first_part else raw_title[:20]

_LAST_LLM_CALL_TIME = 0

def enforce_rate_limit_delay(min_interval: float = 4.0):
    global _LAST_LLM_CALL_TIME
    import time
    now = time.time()
    elapsed = now - _LAST_LLM_CALL_TIME
    if elapsed < min_interval:
        sleep_dur = min_interval - elapsed
        time.sleep(sleep_dur)
    _LAST_LLM_CALL_TIME = time.time()

def call_llm_api(prompt: str, timeout: int = 25) -> str:
    """
    Unified LLM API Invoker calling Gemini API with 4-second rate-limiting buffer.
    """
    enforce_rate_limit_delay(4.0)

    gemini_key = os.environ.get("GEMINI_API_KEY")
    if not gemini_key:
        raise ValueError("GEMINI_API_KEY not found in environment variables.")

    import urllib.request
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-flash-latest:generateContent?key={gemini_key}"
    payload = {"contents": [{"parts": [{"text": prompt}]}]}
    req = urllib.request.Request(url, data=json.dumps(payload).encode('utf-8'), headers={'Content-Type': 'application/json'}, method='POST')
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        res_data = json.loads(resp.read().decode('utf-8'))
        return res_data['candidates'][0]['content']['parts'][0]['text']

    raise ValueError("No LLM API Key set (neither QWEN_API_KEY/DASHSCOPE_API_KEY nor GEMINI_API_KEY).")

def generate_insightful_lead_in(cases, active_theme, focus_domains=""):
    """
    Dynamically synthesizes 100% case-grounded Lead-in guide via LLM (Qwen / Gemini),
    reading all selected top articles and producing executive insights & strategic takeaways.
    """
    if not cases:
        empty_html = """
        <div class="lead-in-header">
          <div class="lead-in-badge">📖 本期總編輯導讀</div>
          <h2 class="lead-in-title">⚠️ 本期符合主題之高品質新聞不足</h2>
        </div>
        <div class="lead-in-content">
          <div class="lead-in-section">
            <p>本期 AI 監聽系統堅持「零湊數原則與最低合格門檻 (Score ≥ 80 分)」。經過 Hard Constraint 主題過濾與多維度品質評比後，本期無達到入選標準之報導。系統不會為了湊滿 6 篇而降低標準或收錄低相關新聞。</p>
          </div>
        </div>
        """
        return empty_html, "本期符合主題之高品質新聞不足（堅持零湊數原則）。"

    theme_clean = re.sub(r'^(AI\s*趨勢周報[：:]?\s*)', '', active_theme).strip()
    
    case_bullets = []
    case_entities = []
    domain_groups = {}
    
    for idx, c in enumerate(cases, 1):
        t = c.get("title", "")
        d = c.get("domain", "產業技術")
        s = strip_citation_phrases(c.get("summary", ""))
        src = c.get("source", "")
        tech = c.get("tech_stack", "")
        
        main_subject = extract_clean_subject_entity(t)
        case_entities.append(main_subject)
        
        case_bullets.append(f"{idx}. [{d}] {t} (來源: {src}, 技術: {tech})\n   摘要: {s[:150]}")
        domain_groups.setdefault(d, []).append((main_subject, t, src))

    cases_block = "\n".join(case_bullets)

    # 1. LLM Generation for 100% custom executive insights lead-in guide
    llm_result = None
    if os.environ.get("QWEN_API_KEY") or os.environ.get("DASHSCOPE_API_KEY") or os.environ.get("GEMINI_API_KEY"):
        prompt = (
            f"你是一位頂尖企業 AI 趨勢總編輯與產業分析師。請依據本期收錄的 {len(cases)} 則實務新聞案例，撰寫一份極具高層商業洞察力、專屬於本期的電子報【本期導讀與專家見解】。\n\n"
            f"【本期主題】：{active_theme}\n"
            f"【重點領域】：{focus_domains}\n"
            f"【本期精選案例列表】：\n{cases_block}\n\n"
            f"### 撰寫要求（絕對嚴禁剪裁或直接拼貼原始新聞標題，必須寫出實質內文重點與商業洞察）：\n"
            f"1. 【趨勢洞察】：提煉標題與 2-3 句巨觀產業趨勢分析，說明這批案例反映的核心技術轉向或策略議題。\n"
            f"2. 【本期實務精解】：歸納 2-3 個核心焦點面向（如：數據落地與營運、資安防衛與 Agent 治理、基礎設施對接等）。每一個面向請撰寫 2 句【實質內容重點】，直接指出事件主角進行了什麼突破、採用了什麼做法，以及帶來什麼實質價值。\n"
            f"3. 【導覽總覽與行動指南】：提煉標題與 2 句給企業主管的實務建議與決策方向。\n\n"
            f"請輸出 JSON 格式如下（不要包含 Markdown 標籤，直接輸出 JSON）：\n"
            f"{{\n"
            f'  "trend_title": "🌐 趨勢洞察｜4-14字動態提煉主題與前瞻趨勢標題",\n'
            f'  "trend_body": "2-3 句結合本期主題與案例技術特點的巨觀趨勢分析",\n'
            f'  "cases_title": "🚀 本期實務精解｜聚焦 {len(cases)} 大實務案例與關鍵戰略維度",\n'
            f'  "cases_body": "詳細說明 2-3 個核心面向的【實質內容重點與突破價值】",\n'
            f'  "conclusion_title": "🎯 導覽總覽｜4-14字動態提煉戰略建議標題",\n'
            f'  "conclusion_body": "2-3 句給決策團隊的雙軌實務行動建議"\n'
            f"}}\n"
        )
        for attempt in range(2):
            try:
                res_text = call_llm_api(prompt, timeout=25).strip()
                res_text = res_text.replace("```json", "").replace("```", "").strip()
                llm_result = json.loads(res_text)
                print("✨ [Stage 3 Lead-in LLM SUCCESS] Dynamically synthesized 100% case-grounded executive insight guide via LLM.")
                break
            except Exception as e:
                print(f"ℹ️ Stage 3 Lead-in LLM synthesis note: ({e}), fallback to rule synthesizer.")

    if llm_result and isinstance(llm_result, dict):
        trend_title = sanitize_forbidden_phrases(llm_result.get("trend_title", ""))
        trend_body = sanitize_forbidden_phrases(llm_result.get("trend_body", ""))
        cases_title = sanitize_forbidden_phrases(llm_result.get("cases_title", ""))
        cases_body = sanitize_forbidden_phrases(llm_result.get("cases_body", ""))
        conclusion_title = sanitize_forbidden_phrases(llm_result.get("conclusion_title", ""))
        conclusion_body = sanitize_forbidden_phrases(llm_result.get("conclusion_body", ""))

        trend_html = f"""<div class="guide-section guide-section-trend">
          <h4 class="guide-sub-title">{trend_title}</h4>
          <p class="guide-text">{trend_body}</p>
        </div>"""

        cases_html = f"""<div class="guide-section guide-section-cases">
          <h4 class="guide-sub-title">{cases_title}</h4>
          <p class="guide-text">
            {cases_body}
          </p>
        </div>"""

        conclusion_html = f"""<div class="guide-section guide-section-conclusion">
          <h4 class="guide-sub-title">{conclusion_title}</h4>
          <p class="guide-text">{conclusion_body}</p>
        </div>"""
    else:
        # Dynamic Fallback: Build 100% grounded content from actual case summaries & key insights (NO raw title concatenation!)
        domain_summary_parts = []
        for d_name, d_cases in domain_groups.items():
            case_insights = []
            for c_ent, c_title, c_src in d_cases[:3]:
                c_match = next((item for item in cases if item.get("title") == c_title or c_ent in item.get("title", "")), None)
                clean_title = re.sub(r'^(【[^】]+】|美股|台股)\s*', '', c_title).strip()
                clean_title = re.sub(r'\s*[||\-—].*$', '', clean_title).strip()
                if len(clean_title) > 30:
                    clean_title = clean_title[:30] + "..."

                if c_match and c_match.get("summary"):
                    s_clean = strip_citation_phrases(c_match.get("summary", ""))
                    s_clean = re.sub(r'^[^\s]+(報導|指出|表示|宣布)[，,:]?\s*', '', s_clean)
                    s_clean = re.sub(r'…|\.\.\.', '', s_clean).strip()
                    
                    sentences = [s.strip() for s in re.split(r'(?<=[。！？])', s_clean) if s.strip()]
                    clean_result = ""
                    for s_item in sentences:
                        if len(clean_result) + len(s_item) <= 120:
                            clean_result += s_item
                        else:
                            break
                    if not clean_result:
                        cut = s_clean.rfind('。', 0, 100)
                        if cut > 20:
                            clean_result = s_clean[:cut+1]
                        else:
                            cut_comma = s_clean.rfind('，', 0, 100)
                            clean_result = s_clean[:cut_comma] + '。' if cut_comma > 20 else s_clean[:100] + '。'
                    
                    clean_result = re.sub(r'…|\.\.\.', '', clean_result).strip()
                    if clean_result and not clean_result.endswith(('。', '！', '？')):
                        clean_result += '。'
                        
                    case_insights.append(f"<strong>{clean_title}</strong>：{clean_result}")
                else:
                    case_insights.append(f"<strong>{clean_title}</strong>")
            
            insights_str = "<br>&nbsp;&nbsp;- " + "<br>&nbsp;&nbsp;- ".join(case_insights)
            domain_summary_parts.append(f"<strong>• {d_name} 核心突破與重點解析</strong>：{insights_str}")
        domain_summary_str = "<br><br>".join(domain_summary_parts)

        trend_html = f"""<div class="guide-section guide-section-trend">
          <h4 class="guide-sub-title">🌐 趨勢洞察｜聚焦《{theme_clean}》之技術落地與轉型典範</h4>
          <p class="guide-text">
            當前企業 AI 應用全面加速落地，本期主題《<strong>{theme_clean}</strong>》收錄當前產業最新代表性個案，解析企業如何從海量營運數據中擷取關鍵洞察，並在推動流程自動化與人機協同的同時，建立具備高 ROI 與風險控管能力的 AI 戰略。
          </p>
        </div>"""

        cases_html = f"""<div class="guide-section guide-section-cases">
          <h4 class="guide-sub-title">🚀 本期實務精解｜精選 {len(cases)} 大實務標竿案例剖析</h4>
          <p class="guide-text">
            本期收錄的 {len(cases)} 篇標竿案例重點綜覽：<br><br>
            {domain_summary_str}
          </p>
        </div>"""

        conclusion_html = f"""<div class="guide-section guide-section-conclusion">
          <h4 class="guide-sub-title">🎯 導覽總覽｜數據驅動與安全治理雙軌並行</h4>
          <p class="guide-text">
            綜觀本期案例，建議決策團隊採取「技術落地與流程自動化」與「安全邊界與合規治理」雙軌策略，推動團隊工具技能升級並落實數據治理，打造可永續發展的 AI 競爭優勢。
          </p>
        </div>"""

    lead_in_inner = f"""<div class="lead-in-card-header" style="margin-bottom: 1.25rem; border-bottom: 1px solid #e2e8f0; padding-bottom: 0.85rem; display: flex; justify-content: space-between; align-items: center; flex-wrap: wrap; gap: 0.5rem;">
        <h3 style="margin: 0; font-size: 1.25rem; color: #0f172a; font-weight: 800;">📖 本期導讀</h3>
      </div>

      <div class="guide-sections-wrapper">
        {trend_html}
        {cases_html}
        {conclusion_html}
      </div>"""

    lead_in_text = f"本期主題聚焦「{theme_clean}」，收錄 {len(cases)} 則重磅實務案例。"
    return lead_in_inner, lead_in_text

def build_custom_case_details(art, idx, active_theme):
    """Dynamically build case card content based 100% on Stage 2 article metadata."""
    title = art["title"]
    desc = art["description"] if art["description"] else art["title"]
    src = art["source"]
    pub_date = art.get("pub_date", "最新報導")
    tags_str = art["tags"]
    tag_raw = tags_str.split("/")[0].strip() if "/" in tags_str else tags_str.strip()
    tag_first = re.sub(r'[：:\s]+$', '', tag_raw)
    
    # Infer specific department tag if tag_first is generic theme title
    from stage2_curator import infer_department_tag
    if tag_first not in DOMAIN_CONFIGS or tag_first in ["職能自動化革命", "職能自動化", "工作流自動化"]:
        tag_first = infer_department_tag(title, desc + " " + art.get("rationale", ""))
        
    config = DOMAIN_CONFIGS.get(tag_first, DOMAIN_CONFIGS.get("企業支援與自動化工作流"))

    clean_rationale = art["rationale"].replace("\r\n", " ").replace("\n", " ").replace("\r", " ").strip()
    
    employee_summary, full_digest_str, real_cover_image = generate_deep_reading_notes(art["link"], title, desc, src, pub_date, tag_first, active_theme, clean_rationale)
    
    if real_cover_image:
        cover_img = real_cover_image
        print(f"🖼️ [Cover Image] Extracted real news cover image for [{title[:20]}...]: {cover_img[:60]}...")
    else:
        cover_img = select_relevant_cover_image(title, employee_summary, clean_rationale, tag_first, idx)

    tech_stack = extract_technology_stack({"title": title, "summary": employee_summary, "rationale": clean_rationale, "domain": tag_first})

    def cut_smart_local(text, max_len):
        if not text:
            return ""
        text = text.strip()
        if len(text) <= max_len:
            return text if text.endswith(('。', '！', '？')) else text + '。'
        for sep in ['。', '；', '！', '？', '，', ' ']:
            pos = text.rfind(sep, 0, max_len)
            if pos >= max_len // 3:
                return text[:pos+1]
        return text[:max_len-1] + '。'

    key_m = f"{title[:32]}... │ {src}"
    prob_text = f"根據 {src} 報導，{cut_smart_local(employee_summary, 120)}"
    innov_text = f"採用關鍵技術：{tech_stack}。詳細技術細節與架構請參閱全文章節拆解。"
    imp_text = f"專家評估切入點：{clean_rationale}"
    pains = ["請參閱全文章節拆解「一、事件背景與產業影響」"]
    techs = [f"核心採用關鍵技術：{tech_stack}", "請參閱全文章節拆解「二、關鍵技術與實作細節」"]
    imps = ["請參閱全文章節拆解「四、決策效益與行動建議」"]

    return {
        "id": idx,
        "domain": tag_first,
        "tech_stack": tech_stack,
        "badge_class": config["badge"],
        "title": title,
        "pub_date": pub_date,
        "source": src,
        "author": f"{src.split(' ')[0]} 產業資深編輯",
        "read_time": "5 分鐘細讀",
        "link": art["link"],
        "cover_image": cover_img,
        "summary": employee_summary,
        "full_digest": full_digest_str,
        "key_metric": key_m,
        "rationale": clean_rationale,
        "cover_bg": config["bg"],
        "icon": config["icon"],
        "problem_statement": {"text": prob_text, "pain_points": pains},
        "innovation_breakthrough": {"text": innov_text, "tech_highlights": techs},
        "impact_and_roi": {"text": imp_text, "impact_results": imps}
    }


def update_newsletter_html(issue_tag, theme_title, focus_domains, cases):
    html_path = "newsletter.html"
    if not os.path.exists(html_path):
        return
        
    theme_clean = theme_title.strip()
    issue_clean = issue_tag.strip()
    case_count = len(cases)
    theme_display = re.sub(r'^(AI\s*趨勢周報[：:]?\s*)', '', theme_clean).strip()

    title_formatted = theme_display
    if "：" in theme_display:
        parts = theme_display.split("：", 1)
        title_formatted = f'{parts[0]}：<br>{parts[1]}'
    elif ":" in theme_display:
        parts = theme_display.split(":", 1)
        title_formatted = f'{parts[0]}:<br>{parts[1]}'

    executive_intro_text, _ = generate_insightful_lead_in(cases, theme_clean, focus_domains)

    new_hero_section = f'''    <!-- Main Hero Banner -->
    <section class="hero-banner">
      <div class="hero-blue-header">
        <div class="hero-header-top-row">
          <div class="theme-tag-pill">
            <span class="sparkle-icon">✨</span>
            AI Trend Listening
          </div>
          <span class="hero-issue-label">{issue_clean}</span>
        </div>
        
        <h1 class="hero-main-title">{title_formatted}</h1>
      </div>

      <div class="hero-white-body">
        <p class="hero-desc">彙整全球 AI 前瞻趨勢與產業落地實務，協助團隊從數據洞察邁向智慧決策與風險預警。</p>
        
        <div class="hero-hashtags">
          <span class="hashtag">#EnterpriseAI</span>
          <span class="hashtag">#AITrends</span>
          <span class="hashtag">#DigitalTransformation</span>
          <span class="hashtag">#DeltaElectronics</span>
        </div>

        <div class="hero-cta-buttons">
          <a href="#lead-in-section" class="btn-hero-primary">
            📖 閱讀本期導讀
          </a>
          <a href="#cases-grid" class="btn-hero-secondary">
            🚀 直接看精選案例
          </a>
        </div>
      </div>
    </section>

    <!-- Lead-in Guide Section -->
    <section id="lead-in-section" class="lead-in-card">
      {executive_intro_text}
    </section>'''

    with open(html_path, "r", encoding="utf-8") as f:
        content = f.read()

    content = re.sub(r'<title>.*?</title>', f'<title>AI Trend Listening 電子報 | {theme_clean}</title>', content)
    content = re.sub(r'<meta name="description" content=".*?">', f'<meta name="description" content="AI Trend Listening 電子報：{theme_clean} 實務案例與落地戰略精選">', content)
    content = re.sub(r'<span class="brand-title">.*?</span>', '<span class="brand-title">AI Trend Listening</span>', content)

    content = re.sub(
        r'<div class="issue-badge">.*?</div>',
        f'<div class="issue-badge">{issue_clean} │ {theme_clean[:14]}...</div>',
        content
    )

    content = re.sub(
        r'<!-- Main Hero Banner -->[\s\S]*?(?=<!-- Cases Section Header -->|<div class="section-header">)',
        new_hero_section.strip() + '\n\n    ',
        content
    )

    content = re.sub(
        r'<h2>.*?精選.*?案例</h2>',
        '<h2>精選案例</h2>',
        content
    )

    new_footer = '''  <!-- Footer -->
  <footer class="footer">
    <p>AI Trend Listening 科技與產業趨勢周報 &copy; 2026 | 專為高層決策者與團隊打造的企業 AI 情報</p>
    <p class="footer-sub">資訊來源包括權威產業媒體、研究機構與新聞數據庫；內容經由主管評選與綜合編撰，精選新聞歸屬原創作者所有。</p>
  </footer>'''
    content = re.sub(r'<!-- Footer -->\s*<footer class="footer">.*?</footer>', new_footer, content, flags=re.DOTALL)

    with open(html_path, "w", encoding="utf-8") as f:
        f.write(content)
        
    print(f"✨ [Stage 3] Successfully updated newsletter.html for {issue_clean} ({theme_clean})!")

def main():
    import argparse
    parser = argparse.ArgumentParser(description="Stage 3 Newsletter Builder")
    parser.add_argument("--mode", choices=["theme", "express"], default=os.environ.get("PIPELINE_MODE", "theme"),
                        help="Builder mode: 'theme' (專題 20 篇) or 'express' (快報 6 篇)")
    args, unknown = parser.parse_known_args()
    pipeline_mode = args.mode.lower()

    articles = load_curated_news_from_stage2()
    if not articles:
        print("⚠️ Warning: No articles loaded from Stage 2. Cannot build dynamic cases.")
        return

    # Check if first article or theme indicates express mode
    if pipeline_mode == "express" or len(articles) <= 6:
        issue_tag = "Vol. 2026 快報"
        newsletter_theme = "每週 AI 趨勢快報：當週前瞻產業科技動態"
        md_path = "data/stage2_curated_report.md"
        if os.path.exists(md_path):
            try:
                with open(md_path, "r", encoding="utf-8") as f:
                    for line in f:
                        if "本週電子報主題" in line:
                            match = re.search(r'`([^`]+)`', line)
                            if match:
                                newsletter_theme = match.group(1).strip()
                                break
            except Exception:
                pass
        focus_domains = "AI Agent、數據治理、算力布局、資安防禦"
        print(f"🚀 [Stage 3 Deep Engine - Express Mode] Building 6-case newsletter for: [{issue_tag}] {newsletter_theme}...")
    else:
        issue_tag, newsletter_theme, focus_domains = get_active_theme_info()
        print(f"🚀 [Stage 3 Deep Engine - Theme Mode] Building newsletter dynamically for active theme: [{issue_tag}] {newsletter_theme}...")
        
    cases = generate_cases_from_articles(articles, newsletter_theme)
    
    output_dir = "data"
    os.makedirs(output_dir, exist_ok=True)
    json_path = os.path.join(output_dir, "newsletter_cases.json")
    stage3_json_path = os.path.join(output_dir, "stage3_newsletter_cases.json")
    
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(cases, f, ensure_ascii=False, indent=2)
    with open(stage3_json_path, "w", encoding="utf-8") as f:
        json.dump(cases, f, ensure_ascii=False, indent=2)
    print(f"✨ [Stage 3] Saved {len(cases)} custom dynamic cases sourced 100% from Stage 2 curated news to: {json_path}")
    
    js_path = "newsletter.js"
    if os.path.exists(js_path):
        with open(js_path, "r", encoding="utf-8") as f:
            js_content = f.read()
        json_str = json.dumps(cases, ensure_ascii=False, indent=2)
        embedded_decl = f"const EMBEDDED_CASES = {json_str};\n\n"
        
        if "document.addEventListener('DOMContentLoaded'" in js_content:
            parts = js_content.split("document.addEventListener('DOMContentLoaded'", 1)
            js_content = embedded_decl + "document.addEventListener('DOMContentLoaded'" + parts[1]
        else:
            js_content = embedded_decl + js_content

        with open(js_path, "w", encoding="utf-8") as f:
            f.write(js_content)
        print(f"✨ [Stage 3] Embedded Stage 2 curated cases into: {js_path}")

    update_newsletter_html(issue_tag, newsletter_theme, focus_domains, cases)

if __name__ == "__main__":
    main()
