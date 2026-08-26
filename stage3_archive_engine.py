#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AI Trend Listening - Stage 3 Archive Engine: Automatically archives each generated issue into a dedicated folder under archives/
Stores Issue-specific newsletter.html, stage2_curated_news.xlsx, reports, cases, and assets.
"""

import sys
import os
import shutil
import re
import json
import openpyxl

# Force UTF-8 encoding for stdout on Windows
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

def sanitize_filename(name):
    """Sanitize string to be safe for Windows folder names."""
    clean = re.sub(r'[\\/:*?"<>|\s]+', '_', name)
    return clean.strip('_')

def archive_current_issue():
    print("\n==================================================")
    print("📦 [Stage 3 Archive Engine] Archiving current newsletter issue...")
    print("==================================================")
    
    excel_path = "data/weekly_newsletter_theme.xlsx"
    issue_folder = "Issue_30_2026-08-10_智慧製造與供應鏈"
    
    if os.path.exists(excel_path):
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            ws = wb["每周電子報主題設定"]
            for row in range(5, ws.max_row + 1):
                row_vals = [str(ws.cell(row=row, column=c).value or "").strip() for c in range(1, max(ws.max_column + 1, 10))]
                if any("啟用中" in v or "Active" in v for v in row_vals):
                    issue_raw = row_vals[0] or "Vol. 2026 Issue #30"
                    date_raw = row_vals[1] or "2026-08-10"
                    theme_raw = row_vals[2] or "智慧製造與供應鏈"

                    m = re.search(r'#(\d+)', issue_raw)
                    num = m.group(1) if m else "30"

                    clean_date_m = re.search(r'\d{4}-\d{2}-\d{2}', date_raw)
                    clean_date_str = clean_date_m.group(0) if clean_date_m else sanitize_filename(date_raw[:10])
                    clean_theme = sanitize_filename(theme_raw[:16])
                    issue_folder = f"Issue_{num}_{clean_date_str}_{clean_theme}"
                    break
        except Exception as e:
            print(f"⚠️ Notice reading theme Excel: {e}")
            
    archive_dir = os.path.join("archives", issue_folder)
    
    if os.path.exists(archive_dir):
        try:
            shutil.rmtree(archive_dir, ignore_errors=True)
        except Exception:
            pass
            
    os.makedirs(archive_dir, exist_ok=True)
    print(f"📁 Target Archive Folder: {archive_dir}")
    
    stage2_xlsx = os.path.join("data", "stage2_curated_news.xlsx")
    if not os.path.exists(stage2_xlsx) and os.path.exists(os.path.join("data", "agent_b_curated_news.xlsx")):
        stage2_xlsx = os.path.join("data", "agent_b_curated_news.xlsx")

    stage2_md = os.path.join("data", "stage2_curated_report.md")
    if not os.path.exists(stage2_md) and os.path.exists(os.path.join("data", "agent_b_curated_report.md")):
        stage2_md = os.path.join("data", "agent_b_curated_report.md")

    files_to_copy = [
        ("newsletter.html", "newsletter.html"),
        ("newsletter.css", "newsletter.css"),
        ("newsletter.js", "newsletter.js"),
        (stage2_xlsx, "stage2_curated_news.xlsx"),
        (stage2_md, "stage2_curated_report.md"),
        (os.path.join("data", "newsletter_cases.json"), "newsletter_cases.json"),
    ]
    
    data_dir = "data"
    if os.path.exists(data_dir):
        for img in os.listdir(data_dir):
            if img.startswith("case") and img.endswith(".png"):
                files_to_copy.append((os.path.join(data_dir, img), img))
                
    copied_count = 0
    for src, dest_name in files_to_copy:
        if os.path.exists(src):
            dest_path = os.path.join(archive_dir, dest_name)
            try:
                shutil.copy2(src, dest_path)
                print(f"  └─ 📄 Copied: {dest_name}")
                copied_count += 1
            except Exception as e:
                try:
                    # Fallback for Windows file lock (e.g. opened in Excel)
                    with open(src, 'rb') as rf:
                        data = rf.read()
                    with open(dest_path, 'wb') as wf:
                        wf.write(data)
                    print(f"  └─ 📄 Copied (via binary stream): {dest_name}")
                    copied_count += 1
                except Exception as e2:
                    print(f"  └─ ⚠️ Could not copy {dest_name}: {e2}")
            
    try:
        standalone_path = os.path.join(archive_dir, "standalone_newsletter.html")
        with open("newsletter.html", "r", encoding="utf-8") as f:
            html_code = f.read()
        with open("newsletter.css", "r", encoding="utf-8") as f:
            css_code = f.read()
        with open("newsletter.js", "r", encoding="utf-8") as f:
            js_code = f.read()
            
        html_code = html_code.replace('<link rel="stylesheet" href="newsletter.css">', f'<style>\n{css_code}\n</style>')
        html_code = html_code.replace('</body>', f'<script>\n{js_code}\n</script>\n</body>')
        
        with open(standalone_path, "w", encoding="utf-8") as f:
            f.write(html_code)
        print(f"  └─ 🌐 Created Standalone Single-File HTML: standalone_newsletter.html")
    except Exception as e:
        print(f"⚠️ Standalone HTML note: {e}")
        
    print(f"🎉 [Archive SUCCESS] Archived {copied_count} files into: {archive_dir}")
    return archive_dir

if __name__ == "__main__":
    archive_current_issue()
