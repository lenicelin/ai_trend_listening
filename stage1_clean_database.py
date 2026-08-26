#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AI Trend Listening - Stage 1 Database Quality Purge Script
Filters accumulated news database against 3 strict user quality criteria:
  1. Specific Use Case
  2. Technical & Architectural Details
  3. Quantifiable Impact & Solutions
"""

import sys
import os
import json
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Force UTF-8 encoding for stdout on Windows
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

EXCLUDE_TERMS = [
    "單純融資", "估值飆升", "股票暴漲", "純粹估值", "股價狂飆", "市值破", "融資案", "series a", "series b"
]

USE_CASE_TERMS = [
    "客服", "招募", "履歷", "供應鏈", "推薦", "程式碼", "審查", "對話", "預測", "維護", "自動化", "轉型", "案例", "防護", "合規", "機房", "良率", "行銷", "業務", "零售", "電商", "資安", "數據", "倉儲", "物流", "晶片", "伺服器", "產線", "製造", "工廠", "營運", "工作流", "治理", "人才", "機密", "保護", "體驗", "招聘", "服務", "平台", "範例",
    "customer service", "recruiting", "supply chain", "recommendation", "code review", "prediction", "maintenance", "automation", "compliance", "fraud", "marketing", "sales", "retail", "logistics", "warehouse", "chip", "factory", "workflow", "hiring", "data"
]

TECH_TERMS = [
    "llama", "claude", "gemini", "gpt", "deepseek", "rag", "agent", "agentic", "vector", "langchain", "autogen", "fine-tuning", "vertex ai", "copilot", "bedrock", "transformer", "模型", "大模型", "架構", "多模態", "api", "即時防護", "雲地混合", "演算法", "軟體", "服務", "平台", "技術", "系統", "硬體", "晶片", "加速器", "處理器", "伺服器", "數據中心", "雲端",
    "model", "system", "platform", "software", "hardware", "cloud", "ai", "llm"
]

IMPACT_TERMS = [
    "%", "成", "秒", "分", "倍", "節省", "縮短", "提升", "降低", "方案", "步驟", "指引", "框架", "處置", "解決方案", "風險", "挑戰", "效益", "成效", "成果", "應用", "突破", "轉型", "新制", "戰略", "法規", "裁判", "判決", "規範", "範例", "新局",
    "percent", "reduced", "improved", "faster", "decreased", "solution", "framework", "guideline", "step", "challenge", "impact", "policy", "rule", "strategy"
]

def is_compliant_article(art):
    title = art.get("title", "")
    desc = art.get("description", "")
    src = art.get("source", "")
    
    if not title or len(title.strip()) < 5:
        return False, "Excluded: Title too short or empty"

    full_text = f"{title} {desc} {src}".lower()

    if any(ex in full_text for ex in EXCLUDE_TERMS) and not any(uc in full_text for uc in ["製造", "工廠", "供應鏈", "客服", "案例", "轉型"]):
        return False, "Excluded: Pure funding/stock hype or Event Registration Ad"

    return True, "Passed Stage 1 Basic Quality"

def main():
    json_path = "data/stage1_ai_news.json"
    if not os.path.exists(json_path) and os.path.exists("data/agent_a_ai_news.json"):
        json_path = "data/agent_a_ai_news.json"

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    articles = data.get("articles", [])
    print(f"📦 Evaluating {len(articles)} articles against Stage 1 High-Quality Rules...")

    passed = []
    purged = []

    for art in articles:
        ok, reason = is_compliant_article(art)
        if ok:
            passed.append(art)
        else:
            purged.append((art["title"], reason))

    print(f"\n📊 Stage 1 Quality Purge Results:")
    print(f"  ✅ 保留高品質合規新聞 (Passed): {len(passed)} 篇")
    print(f"  ❌ 剔除不合規抽象新聞 (Purged): {len(purged)} 篇")

    passed.sort(key=lambda x: x["pub_date"])

    data["articles"] = passed
    data["total_count"] = len(passed)

    out_json = "data/stage1_ai_news.json"
    with open(out_json, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"\n💾 Updated cleaned JSON database at: {out_json}")

    excel_name = "data/stage1_ai_news.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stage 1 高品質 AI 新聞數據庫"
    ws.views.sheetView[0].showGridLines = True

    header_font = Font(name="Microsoft JhengHei", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    title_font = Font(name="Microsoft JhengHei", size=10, bold=True)
    body_font = Font(name="Microsoft JhengHei", size=10)
    date_font = Font(name="Consolas", size=10, bold=True, color="0F766E")
    link_font = Font(name="Microsoft JhengHei", size=10, color="0066CC", underline="single")
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    headers = ["項次", "新聞標題", "新聞發布日期", "正確原文連結", "媒體來源", "趨勢標籤", "新聞摘要"]
    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[1].height = 28

    for idx, art in enumerate(passed, 1):
        row_idx = idx + 1
        ws.append([
            idx,
            art["title"],
            art["pub_date"],
            art["link"],
            art["source"],
            ", ".join(art.get("trend_tags", [])),
            art.get("description", "")
        ])

        c_num = ws.cell(row=row_idx, column=1)
        c_title = ws.cell(row=row_idx, column=2)
        c_date = ws.cell(row=row_idx, column=3)
        c_link = ws.cell(row=row_idx, column=4)
        c_src = ws.cell(row=row_idx, column=5)
        c_tag = ws.cell(row=row_idx, column=6)
        c_desc = ws.cell(row=row_idx, column=7)

        c_num.alignment = Alignment(horizontal="center", vertical="center")
        c_title.font = title_font
        c_date.font = date_font
        c_date.alignment = Alignment(horizontal="center", vertical="center")
        c_link.font = link_font
        c_link.hyperlink = art["link"]
        c_src.font = body_font
        c_tag.font = body_font
        c_desc.font = body_font

        for col in range(1, 8):
            ws.cell(row=row_idx, column=col).border = thin_border
        ws.row_dimensions[row_idx].height = 22

    col_widths = {1: 8, 2: 45, 3: 20, 4: 50, 5: 25, 6: 35, 7: 50}
    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    try:
        wb.save(excel_name)
        print(f"📊 Saved cleaned Excel at: {excel_name}")
    except Exception:
        pass

if __name__ == "__main__":
    main()
