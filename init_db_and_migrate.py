#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AI Trend Listening - Database Initializer & Migration Utility (PostgreSQL Exclusively)
Initializes database tables in PostgreSQL and migrates datasets from JSON/Excel into PostgreSQL.
"""

import os
import sys
import json
import openpyxl
from db_manager import (
    init_db,
    get_connection,
    batch_upsert_articles,
    save_theme,
    save_curated_articles
)

# Force UTF-8 encoding for stdout on Windows
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

def migrate_stage1_json():
    json_path = os.path.join("data", "stage1_ai_news.json")
    if not os.path.exists(json_path):
        print("ℹ️ No stage1_ai_news.json found to migrate.")
        return 0

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    articles = data if isinstance(data, list) else data.get("articles", [])
    count = batch_upsert_articles(articles)
    print(f"✅ [Migration] Synced {count} Stage 1 articles into active PostgreSQL DB from {json_path}.")
    return count

def migrate_weekly_themes():
    excel_path = os.path.join("data", "weekly_newsletter_theme.xlsx")
    if not os.path.exists(excel_path):
        print("ℹ️ No weekly_newsletter_theme.xlsx found to migrate.")
        return 0

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if "每周電子報主題設定" not in wb.sheetnames:
        return 0

    ws = wb["每周電子報主題設定"]
    count = 0
    for row in range(5, ws.max_row + 1):
        vals = [str(ws.cell(row=row, column=c).value or "").strip() for c in range(1, 10)]
        issue_tag = vals[0]
        issue_date = vals[1]
        theme_title = vals[2]
        focus_domains = vals[3]
        status_raw = vals[4]

        if issue_tag and theme_title:
            status = "Active" if any(k in status_raw for k in ["啟用", "Active"]) else "Completed"
            save_theme(issue_tag, issue_date, theme_title, focus_domains, status=status)
            count += 1

    print(f"✅ [Migration] Synced {count} weekly themes into active PostgreSQL DB from {excel_path}.")
    return count

def migrate_stage2_curated():
    excel_path = os.path.join("data", "stage2_curated_news.xlsx")
    if not os.path.exists(excel_path):
        print("ℹ️ No stage2_curated_news.xlsx found to migrate.")
        return 0

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = None
    for name in ["本期精選新聞清單", "電子報精選新聞列表", "電子報精選新聞清單"]:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        if len(wb.worksheets) >= 2:
            ws = wb.worksheets[1]
        else:
            ws = wb.active

    issue_tag = "Vol. 2026 Issue #4"
    if wb.worksheets and wb.worksheets[0]:
        cell_v = str(wb.worksheets[0].cell(row=2, column=1).value or "")
        if "Issue" in cell_v:
            parts = cell_v.split("：")
            issue_tag = parts[0].replace("本期電子報主題 (", "").replace(")", "").strip()

    curated_list = []
    for row in range(2, ws.max_row + 1):
        idx = ws.cell(row=row, column=1).value
        if not idx:
            continue
        curated_list.append({
            "score": str(ws.cell(row=row, column=2).value or "").strip(),
            "functional_tags": str(ws.cell(row=row, column=3).value or "").strip(),
            "title": str(ws.cell(row=row, column=4).value or "").strip(),
            "pub_date": str(ws.cell(row=row, column=5).value or "").strip(),
            "link": str(ws.cell(row=row, column=6).value or "").strip(),
            "source": str(ws.cell(row=row, column=7).value or "").strip(),
            "rationale": str(ws.cell(row=row, column=8).value or "").strip(),
            "description": str(ws.cell(row=row, column=9).value or "").strip(),
        })

    count = save_curated_articles(issue_tag, curated_list)
    print(f"✅ [Migration] Synced {count} curated articles for [{issue_tag}] into PostgreSQL DB from {excel_path}.")
    return count

def main():
    print("🚀 [DB Migration] Initializing PostgreSQL database...")
    init_db()
    backend, _ = get_connection()
    print(f"ℹ️ Active Database Engine: {backend.upper()}")

    c1 = migrate_stage1_json()
    c2 = migrate_weekly_themes()
    c3 = migrate_stage2_curated()

    print("\n🎉 [DB Migration COMPLETE] PostgreSQL database tables and indexes successfully synced!")

if __name__ == "__main__":
    main()
