#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AI Trend Listening - Stage 1: News Fetcher & Basic Quality Filter (Maximize Recall)
Broadly crawls and accumulates AI & tech news from RSS feeds and Google News:
  - Date limit (from START_DATE onwards)
  - URL / Title / Hash deduplication
  - Basic quality filtering (minimum length, anti-spam)
  - Obvious irrelevant exclusion (pure funding, stock speculation, event/course registration ads)
Exports clean dataset to data/stage1_ai_news.json, csv, and xlsx.
"""

import sys
import os
import re
import csv
import json
import html
import ssl
import email.utils
import urllib.request
import urllib.parse
import xml.etree.ElementTree as ET
from datetime import datetime

# Force UTF-8 encoding for stdout on Windows
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

# Start Date Filter (Accumulate news published from 2026-07-01 onwards)
START_DATE = "2026-07-01"

DATE_WINDOWS = [
    "after:2026-06-30 before:2026-07-08",
    "after:2026-07-07 before:2026-07-15",
    "after:2026-07-14 before:2026-07-22",
    "after:2026-07-21 before:2026-07-27",
    "after:2026-07-26"
]

BASE_DOMAIN_TOPICS = [
    ("TW AI General", "AI 人工智慧"),
    ("US AI General", "Artificial Intelligence"),
    ("TW Manufacturing & Supply Chain", "AI 製造 工廠 供應鏈 物流 半導體"),
    ("US Manufacturing & Supply Chain", "AI manufacturing supply chain factory semiconductor"),
    ("TW R&D & Science", "AI 研發 創新 材料 生物"),
    ("US R&D & Science", "AI research discovery science material design"),
    ("TW Enterprise & HR & Legal", "AI 企業 轉型 招募 人力 合規 資安 法規"),
    ("US Enterprise & HR & Legal", "AI enterprise automation hiring compliance security policy")
]

DOMAIN_QUERIES = []
for label_base, term in BASE_DOMAIN_TOPICS:
    for idx, dw in enumerate(DATE_WINDOWS, 1):
        DOMAIN_QUERIES.append((f"{label_base} (W{idx})", f"{term} {dw}"))

FIXED_FEEDS = [
    # 國際頂級科技與 AI 媒體 RSS
    {"source": "TechCrunch AI", "url": "https://techcrunch.com/category/artificial-intelligence/feed/", "lang": "en"},
    {"source": "Wired AI", "url": "https://www.wired.com/feed/tag/ai/latest/rss", "lang": "en"},
    {"source": "VentureBeat AI", "url": "https://venturebeat.com/category/ai/feed/", "lang": "en"},
    {"source": "Ars Technica AI", "url": "https://feeds.arstechnica.com/arstechnica/index", "lang": "en"},
    {"source": "MIT Technology Review", "url": "https://www.technologyreview.com/feed/", "lang": "en"},
    {"source": "The Verge AI", "url": "https://www.theverge.com/rss/ai-artificial-intelligence/index.xml", "lang": "en"},

    # 台灣主流科技與商業媒體 RSS
    {"source": "iThome 科技報", "url": "https://www.ithome.com.tw/rss", "lang": "zh"},
    {"source": "科技新報 TechNews", "url": "https://technews.tw/feed/", "lang": "zh"},
    {"source": "Inside 硬塞網路趨勢", "url": "https://www.inside.com.tw/feed/rss", "lang": "zh"},
    {"source": "TechOrange 科技報橘", "url": "https://techorange.com/feed/", "lang": "zh"},
    {"source": "經理人月刊", "url": "https://www.managertoday.com.tw/rss", "lang": "zh"},

    # 官方頂尖 AI 實驗室與巨頭 RSS
    {"source": "OpenAI News", "url": "https://openai.com/news/rss.xml", "lang": "en"},
    {"source": "Google DeepMind Blog", "url": "https://deepmind.google/blog/rss.xml", "lang": "en"},
    {"source": "Google Research Blog", "url": "https://research.google/blog/rss/", "lang": "en"},
    {"source": "Google Blog", "url": "https://blog.google/rss/", "lang": "en"},
    {"source": "NVIDIA Newsroom", "url": "https://nvidianews.nvidia.com/releases.xml", "lang": "en"},
    {"source": "AWS Machine Learning Blog", "url": "https://aws.amazon.com/blogs/machine-learning/feed/", "lang": "en"},
    {"source": "Hugging Face Blog", "url": "https://huggingface.co/blog/feed.xml", "lang": "en"}
]

TREND_TAXONOMY = {
    "智慧製造與工業 AI": {
        "keywords": ["製造", "工廠", "工業", "半導體", "生產線", "良率", "設備", "自動化", "數值", "工業4.0", "算力", "晶片", "伺服器", "機房", "factory", "manufacturing", "semiconductor", "yield", "chip"],
        "badge_color": "emerald"
    },
    "供應鏈韌性與物流自動化": {
        "keywords": ["供應鏈", "物流", "倉儲", "運籌", "庫存", "航運", "貨運", "採購", "關稅", "supply chain", "logistics", "warehouse", "inventory", "shipping"],
        "badge_color": "cyan"
    },
    "研發與創新": {
        "keywords": ["研發", "創新", "科學", "材料", "生物", "醫藥", "設計", "模型", "實驗", "論文", "ai for science", "research", "discovery", "science", "r&d"],
        "badge_color": "blue"
    },
    "Agentic AI / 代理式 AI": {
        "keywords": ["agent", "agentic", "代理", "智能體", "多智能體", "autonomous agent", "workflow automation", "copilot", "auto-gpt"],
        "badge_color": "cyan"
    },
    "LLMs & Reasoning / 大語言模型與推理": {
        "keywords": ["llm", "reasoning", "gpt", "claude", "gemini", "llama", "deepseek", "transformer", "推理", "大模型", "語言模型", "prompt", "o1", "o3"],
        "badge_color": "emerald"
    },
    "Sovereign AI & Policy / 主權 AI 與法規": {
        "keywords": ["sovereign", "regulation", "policy", "act", "eu ai act", "copyright", "governance", "主權", "法規", "監管", "著作權", "隱私", "安全", "safety", "資安", "合規"],
        "badge_color": "amber"
    },
    "Chips & Hardware / 晶片與算力": {
        "keywords": ["nvidia", "amd", "intel", "chip", "gpu", "semiconductor", "tsmc", "台積電", "晶片", "算力", "數據中心", "data center", "h100", "blackwell"],
        "badge_color": "rose"
    },
    "Enterprise & ROI / 企業應用與效益": {
        "keywords": ["enterprise", "business", "roi", "productivity", "cloud", "saas", "企業", "商業", "生產力", "應用", "投資", "revenue", "startup"],
        "badge_color": "purple"
    }
}

STAT_KEYWORDS = [
    "Agentic AI", "OpenAI", "GPT-5", "Claude", "Gemini", "NVIDIA", "台積電", 
    "具身智能", "主權 AI", "EU AI Act", "DeepSeek", "機器人", "晶片", "LLM"
]

EXCLUDE_TERMS = [
    "單純融資", "估值飆升", "股票暴漲", "純粹估值", "股價狂飆", "市值破", "融資案", "series a", "series b",
    "論壇報名", "活動報名", "報名開放", "開放申請", "論壇台北場", "歡迎報名", "線上論壇", "研討會報名", "報名網址", "課程報名", "免費報名", "席次有限", "報名簡章", "報名連結", "d forum", "forum", "企業論壇", "活動平台", "趨勢論壇", "summit", "event go"
]

STAGE1_EXPORT_HEADERS = ["項次", "新聞標題", "新聞發布日期", "正確原文連結", "媒體來源", "趨勢標籤", "新聞摘要"]

def build_stage1_export_row(idx, art):
    """
    Code Review Item 2: Unified Export Row Generator for CSV and Excel.
    Guarantees identical column definitions and formatting across CSV and Excel exports.
    """
    trend_tags_str = ", ".join(art["trend_tags"]) if isinstance(art["trend_tags"], list) else str(art.get("trend_tags", ""))
    return [
        idx,
        art["title"],
        art["pub_date"],
        art["link"],
        art["source"],
        trend_tags_str,
        art["description"]
    ]

def analyze_article_metadata(title, description):
    """
    Code Review Item 3: Single-Pass scanner for Trend Tagging and Keyword Statistics.
    Scans article text ONCE to infer trend tags and track keyword occurrences.
    """
    full_text = f"{title} {description}".lower()
    matched_tags = []
    for cat_name, cat_data in TREND_TAXONOMY.items():
        for kw in cat_data["keywords"]:
            if kw.lower() in full_text:
                matched_tags.append(cat_name)
                break

    if not matched_tags:
        matched_tags = ["Enterprise & ROI / 企業應用與效益"]

    matched_stat_kws = [kw for kw in STAT_KEYWORDS if kw.lower() in full_text]
    return matched_tags, matched_stat_kws

def is_basic_quality_article(title, description, source=""):
    """
    Stage 1 Basic Quality & Anti-Spam Filter (Maximize Recall).
    Filters out empty/too short titles and pure funding, stock speculation, or event/registration ads.
    Does NOT strictly require specific use case, tech, or ROI terms (which are scored in Stage 2).
    """
    if not title or len(title.strip()) < 5:
        return False, "Excluded: Title too short or empty"
        
    full_text = f"{title} {description} {source}".lower()
    
    # Exclude pure funding / stock speculation / event registration ads
    if any(ex in full_text for ex in EXCLUDE_TERMS):
        return False, "Excluded: Pure funding/stock hype or Event Registration Ad"
        
    return True, "Passed Stage 1 Basic Quality"

def clean_text(raw_html):
    if not raw_html:
        return ""
    cleanr = re.compile('<.*?>')
    cleantext = re.sub(cleanr, '', raw_html)
    return html.unescape(cleantext).strip()

def format_pub_date(raw_date_str):
    if not raw_date_str:
        return "未知日期"
    try:
        dt = email.utils.parsedate_to_datetime(raw_date_str)
        if dt:
            return dt.strftime("%Y-%m-%d %H:%M")
    except Exception:
        pass

    match = re.search(r'(\d{4})[-/](\d{1,2})[-/](\d{1,2})[T\s]?(\d{1,2})?:?(\d{1,2})?', raw_date_str)
    if match:
        y, m, d, hh, mm = match.groups()
        hh_str = f" {int(hh):02d}:{int(mm):02d}" if hh and mm else " 00:00"
        return f"{y}-{int(m):02d}-{int(d):02d}{hh_str}"

    return raw_date_str

def parse_rss_item(item, default_source="Google News"):
    """
    Code Review Item 5: RSS Item Parsing helper function.
    Extracts title, link, pubDate, description, and source tag fallbacks cleanly across RSS & Atom formats.
    """
    title_elem = item.find('title')
    title = clean_text(title_elem.text) if title_elem is not None and title_elem.text else ""

    link = ""
    link_elem = item.find('link')
    if link_elem is not None:
        if link_elem.text and link_elem.text.strip():
            link = link_elem.text.strip()
        elif 'href' in link_elem.attrib:
            link = link_elem.attrib['href']

    pub_elem = None
    for tag in ['pubDate', '{http://www.w3.org/2005/Atom}published', '{http://www.w3.org/2005/Atom}updated', 'date']:
        found = item.find(tag)
        if found is not None and found.text and found.text.strip():
            pub_elem = found
            break
    raw_pub_date = pub_elem.text.strip() if pub_elem is not None else ""
    pub_date = format_pub_date(raw_pub_date)

    desc_elem = None
    for tag in ['description', '{http://www.w3.org/2005/Atom}summary', '{http://www.w3.org/2005/Atom}content']:
        found = item.find(tag)
        if found is not None and found.text and found.text.strip():
            desc_elem = found
            break
    description = clean_text(desc_elem.text) if desc_elem is not None else ""

    source_name = default_source
    source_elem = item.find('source')
    if source_elem is not None and source_elem.text:
        source_name = f"{source_elem.text.strip()} (via Google News)"

    return title, link, pub_date, description, source_name

def parse_rss_url(url, default_source="Google News"):
    articles = []
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36'}
    req = urllib.request.Request(url, headers=headers)
    
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE

    try:
        with urllib.request.urlopen(req, timeout=10, context=ctx) as response:
            content = response.read()
            root = ET.fromstring(content)
            
            channel = root.find('channel')
            items = channel.findall('item') if channel is not None else root.findall('{http://www.w3.org/2005/Atom}entry')
            
            for item in items[:40]:
                title, link, pub_date, description, source_name = parse_rss_item(item, default_source)
                
                if not title or not link:
                    continue
                
                # Date Filter: Must be on or after START_DATE
                if START_DATE and len(pub_date) >= 10 and pub_date[:10] < START_DATE:
                    continue

                is_quality, reason = is_basic_quality_article(title, description, source_name)
                if not is_quality:
                    continue
                
                matched_tags, _ = analyze_article_metadata(title, description)

                articles.append({
                    "title": title,
                    "link": link,
                    "pub_date": pub_date,
                    "description": description[:250] + ("..." if len(description) > 250 else ""),
                    "source": source_name,
                    "trend_tags": matched_tags
                })
    except Exception as e:
        print(f"[Warning] Could not fetch RSS from {url[:60]}: {e}")
        
    return articles

def calculate_keyword_stats(articles):
    """Calculates keyword statistics from articles dataset."""
    stats = {kw: 0 for kw in STAT_KEYWORDS}
    for item in articles:
        text = f"{item['title']} {item['description']}".lower()
        for kw in STAT_KEYWORDS:
            if kw.lower() in text:
                stats[kw] += 1
    sorted_stats = dict(sorted(stats.items(), key=lambda x: x[1], reverse=True))
    return sorted_stats

def save_excel_and_csv(output_dir, all_articles):
    csv_path = os.path.join(output_dir, "stage1_ai_news.csv")
    try:
        with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(STAGE1_EXPORT_HEADERS)
            for idx, art in enumerate(all_articles, 1):
                writer.writerow(build_stage1_export_row(idx, art))
        print(f"[Saved] CSV dataset at: {csv_path}")
    except Exception as e:
        print(f"[Notice] CSV export note: {e}")

    xlsx_path = os.path.join(output_dir, "stage1_ai_news.xlsx")
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stage 1 AI 新聞即時監聽與累積數據"
        ws.views.sheetView[0].showGridLines = True

        header_font = Font(name="Microsoft JhengHei", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        title_font = Font(name="Microsoft JhengHei", size=10, bold=True)
        body_font = Font(name="Microsoft JhengHei", size=10)
        date_font = Font(name="Consolas", size=10, bold=True, color="0F766E")
        link_font = Font(name="Microsoft JhengHei", size=10, color="0066CC", underline="single")
        
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )

        ws.append(STAGE1_EXPORT_HEADERS)

        for col_num in range(1, len(STAGE1_EXPORT_HEADERS) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws.row_dimensions[1].height = 28

        for idx, art in enumerate(all_articles, 1):
            row_idx = idx + 1
            ws.append(build_stage1_export_row(idx, art))

            c_num = ws.cell(row=row_idx, column=1)
            c_title = ws.cell(row=row_idx, column=2)
            c_date = ws.cell(row=row_idx, column=3)
            c_link = ws.cell(row=row_idx, column=4)
            c_src = ws.cell(row=row_idx, column=5)
            c_tag = ws.cell(row=row_idx, column=6)
            c_desc = ws.cell(row=row_idx, column=7)

            c_num.alignment = Alignment(horizontal="center", vertical="center")
            c_title.font = title_font
            c_date.font = date_font
            c_date.alignment = Alignment(horizontal="center", vertical="center")
            c_link.font = link_font
            c_link.hyperlink = art["link"]
            c_src.font = body_font
            c_tag.font = body_font
            c_desc.font = body_font

            for col in range(1, 8):
                ws.cell(row=row_idx, column=col).border = thin_border
            ws.row_dimensions[row_idx].height = 22

        col_widths = {1: 8, 2: 45, 3: 20, 4: 50, 5: 25, 6: 35, 7: 50}
        for col_idx, width in col_widths.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

        wb.save(xlsx_path)
        print(f"[Saved] Excel workbook at: {xlsx_path}")
    except PermissionError:
        print(f"⚠️ [Excel Error] Could not update {xlsx_path} because the file is currently locked or open in Microsoft Excel. Please close Excel and re-run.")
    except Exception as e:
        print(f"[Notice] Excel export note: {e}")

def main():
    print(f"📡 [Stage 1 INCREMENTAL CRAWL] Fetching & accumulating AI news starting from [{START_DATE}] to present...")
    output_dir = "data"
    os.makedirs(output_dir, exist_ok=True)
    json_path = os.path.join(output_dir, "stage1_ai_news.json")
    
    all_articles = []
    seen_links = set()
    seen_titles = set()

    # Load existing database to prevent overwriting/deleting past news
    if os.path.exists(json_path):
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                old_data = json.load(f)
                all_articles = old_data.get("articles", [])
                for art in all_articles:
                    seen_links.add(art["link"])
                    seen_titles.add(art["title"].lower().strip())
            print(f"📦 Loaded {len(all_articles)} existing articles from incremental database ({json_path}).")
        except Exception as e:
            print(f"⚠️ [Notice] Could not load existing database, starting clean: {e}")

    initial_count = len(all_articles)
    fetched_new_items = []
    
    # Pure Broad AI Domain Queries (Purely decoupled from weekly active theme)
    search_queries = DOMAIN_QUERIES

    # Code Review Item 4: Optimized Google News RSS fetching by target region/language
    for label, query_str in search_queries:
        enc_q = urllib.parse.quote(query_str)
        if label.startswith("TW"):
            rss_url = f"https://news.google.com/rss/search?q={enc_q}&hl=zh-TW&gl=TW&ceid=TW:zh-Hant"
            fetched_new_items.extend(parse_rss_url(rss_url, default_source=f"Google News TW ({label})"))
        else:
            rss_url = f"https://news.google.com/rss/search?q={enc_q}&hl=en-US&gl=US&ceid=US:en"
            fetched_new_items.extend(parse_rss_url(rss_url, default_source=f"Google News US ({label})"))

    print(f"📡 Fetching from {len(FIXED_FEEDS)} top-tier RSS feeds (from {START_DATE} to present)...")
    for feed in FIXED_FEEDS:
        fetched_new_items.extend(parse_rss_url(feed["url"], default_source=feed["source"]))

    new_added_count = 0
    for item in fetched_new_items:
        t_key = item["title"].lower().strip()
        if item["link"] not in seen_links and t_key not in seen_titles:
            seen_links.add(item["link"])
            seen_titles.add(t_key)
            all_articles.append(item)
            new_added_count += 1

    print(f"✨ [Stage 1 Incremental SUCCESS] Added {new_added_count} new articles. Total accumulated: {len(all_articles)} articles (from {START_DATE} onwards).")

    all_articles.sort(key=lambda x: x["pub_date"])

    keyword_stats = calculate_keyword_stats(all_articles)
    trend_counts = {cat: 0 for cat in TREND_TAXONOMY.keys()}
    for art in all_articles:
        for tag in art["trend_tags"]:
            trend_counts[tag] = trend_counts.get(tag, 0) + 1

    json_data = {
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "total_count": len(all_articles),
        "start_date": START_DATE,
        "trend_counts": trend_counts,
        "keyword_stats": keyword_stats,
        "articles": all_articles
    }
    
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(json_data, f, ensure_ascii=False, indent=2)
    print(f"[Saved] Incremental JSON dataset at: {json_path}")
    
    save_excel_and_csv(output_dir, all_articles)

    try:
        from db_manager import batch_upsert_articles
        db_count = batch_upsert_articles(all_articles)
        print(f"[Saved] Synced {db_count} articles into PostgreSQL Database (Supabase Cloud).")
    except Exception as e:
        print(f"⚠️ [Notice] Could not sync to PostgreSQL DB: {e}")

    print("✨ Stage 1 incremental news crawl completed successfully!")

if __name__ == "__main__":
    main()
