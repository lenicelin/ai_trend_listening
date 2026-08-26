#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Creates data/weekly_newsletter_theme.xlsx for the user to manage weekly newsletter themes.
Updates stage2_curator.py to dynamically read the active theme from this Excel file!
"""

import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Force UTF-8 encoding for stdout on Windows
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

def create_theme_excel():
    output_dir = "data"
    os.makedirs(output_dir, exist_ok=True)
    excel_path = os.path.join(output_dir, "weekly_newsletter_theme.xlsx")
    
    wb = openpyxl.Workbook()
    
    # Sheet 1: 每周主題設定
    ws_theme = wb.active
    ws_theme.title = "每周電子報主題設定"
    ws_theme.views.sheetView[0].showGridLines = True
    
    h_font = Font(name="Microsoft JhengHei", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Royal Blue
    
    title_font = Font(name="Microsoft JhengHei", size=14, bold=True, color="1E3A8A")
    sub_font = Font(name="Microsoft JhengHei", size=10, color="475569")
    
    ws_theme.cell(row=1, column=1, value="📋 Stage 2 每周電子報主題管理清單").font = title_font
    ws_theme.cell(row=2, column=1, value="提示：Stage 2 將自動讀取【啟用狀態】為「啟用中 (Active)」的行次作為本周評比主題！").font = sub_font
    ws_theme.cell(row=3, column=1, value="")
    
    headers = ["期數 (Issue)", "發行日期 (Date)", "電子報本周主題 (Weekly Theme)", "重點對應領域 (Focus Domains)", "選題目標與焦點描述 (Curator Goal & Target)", "啟用狀態 (Status)", "備註與特別說明 (Notes)"]
    ws_theme.append(headers)
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws_theme.cell(row=4, column=col_idx)
        cell.font = h_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws_theme.row_dimensions[4].height = 26
    
    # Sample Theme Rows
    sample_data = [
        ["Vol. 2026 Issue #1", "2026-07-28", "AI 驅動的策略決策：從數據洞察到風險預警", "策略決策, 數據洞察, 風險預警, 商業智慧", "本期希望尋找高層如何運用 AI 儀表板進行營運預警的案例，特別重視有實質 ROI 或營收數據說明的報導。", "已發行", "模式 A (特定主題)：聚焦 AI 策略決策與風險預警案例"],
        ["Vol. 2026 Issue #2", "2026-08-05", "全網趨勢動態探索 (無預設主題)", "動態聚類, 全網熱點, 綜合趨勢, AI關鍵新聞", "【模式 B：不預設主題】自動分析當週全網熱度最高的前 4 大真實趨勢，由系統動態歸納題材並推薦最適標題！", "啟用中 (Active)", "模式 B (全網動態探索)：不設限主題，由 AI 自動探索當週 4 大熱點"],
        ["Vol. 2026 Issue #3", "2026-08-12", "職能自動化革命： AI 新夥伴", "職能自動化, AI協作, 人機協作, 工作流自動化", "重點挑選企業運用 AI Agent 或自動化工具賦能第一線員工（如 HR、財務、供應鏈、資安）之實務落地案例。", "預備中", "模式 A (特定主題)：聚焦職能自動化與人機協同案例"],
        ["Vol. 2026 Issue #4", "2026-08-19", "製造與供應鏈 AI 轉型：智慧工廠與韌性物流", "製造, 供應鏈, 物流, 工業AI, 產線自動化", "聚焦智慧工廠產線 AI 檢測、預測性維護與供應鏈彈性調度的標竿案例。", "預備中", "模式 A (特定主題)：聚焦智慧製造與供應鏈韌性案例"]
    ]
    
    active_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Light emerald
    active_font = Font(name="Microsoft JhengHei", size=10, bold=True, color="166534")
    
    b_font = Font(name="Microsoft JhengHei", size=10)
    date_font = Font(name="Consolas", size=10)
    
    for row_idx, row_data in enumerate(sample_data, start=5):
        ws_theme.append(row_data)
        for c_idx in range(1, len(row_data) + 1):
            cell = ws_theme.cell(row=row_idx, column=c_idx)
            cell.font = b_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if c_idx in [1, 2, 6]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if c_idx == 2:
                cell.font = date_font
                
        if "啟用中" in row_data[5]:
            ws_theme.cell(row=row_idx, column=6).fill = active_fill
            ws_theme.cell(row=row_idx, column=6).font = active_font
            
        ws_theme.row_dimensions[row_idx].height = 24

    col_widths = {1: 20, 2: 16, 3: 45, 4: 30, 5: 55, 6: 18, 7: 40}
    for col_i, w in col_widths.items():
        ws_theme.column_dimensions[get_column_letter(col_i)].width = w

    # Sheet 2: 使用說明
    ws_guide = wb.create_sheet(title="填寫與操作說明")
    ws_guide.views.sheetView[0].showGridLines = True
    
    ws_guide.cell(row=1, column=1, value="📖 Weekly Newsletter Theme Settings Guide").font = title_font
    
    instructions = [
        "1. 如何變更本周主題：只需將您想執行的主題那一行的【啟用狀態】修改為「啟用中 (Active)」，其他列改為「預備中」或「已發行」。",
        "2. 如何新增未來主題：在表格下方直接新增一行，填寫【期數】、【發行日期】、【電子報本周主題】與【重點領域】即可。",
        "3. Stage 2 自動對接：執行 stage2_curator.py 時，系統會自動讀取此 Excel 檔中的【啟用中】主題，據此評比 Stage 1 資料庫中的新聞！",
        "4. 電子報自動更新：Stage 2 完成評比後，Stage 3 會自動依據該主題生成對應的 HTML 電子報。"
    ]
    
    for idx, inst in enumerate(instructions, start=3):
        ws_guide.cell(row=idx, column=1, value=inst).font = b_font
        ws_guide.row_dimensions[idx].height = 22
        
    ws_guide.column_dimensions['A'].width = 100

    wb.save(excel_path)
    print(f"✨ Successfully generated theme Excel manager at: {excel_path}")

if __name__ == "__main__":
    create_theme_excel()
