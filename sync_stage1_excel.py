import sys
import os
import csv
import json

# Force UTF-8 stdout
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def sync_excel_from_csv():
    csv_path = os.path.join("data", "stage1_ai_news.csv")
    json_path = os.path.join("data", "stage1_ai_news.json")
    xlsx_path = os.path.join("data", "stage1_ai_news.xlsx")

    articles = []

    # Read from CSV to ensure 100% exact match
    if os.path.exists(csv_path):
        with open(csv_path, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            header = next(reader, None)
            for row in reader:
                if len(row) >= 7:
                    articles.append({
                        "idx": row[0],
                        "title": row[1],
                        "pub_date": row[2],
                        "link": row[3],
                        "source": row[4],
                        "trend_tags": row[5],
                        "description": row[6]
                    })
        print(f"📥 Loaded {len(articles)} rows from CSV ({csv_path}).")
    elif os.path.exists(json_path):
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            arts = data.get("articles", [])
            for idx, art in enumerate(arts, 1):
                articles.append({
                    "idx": idx,
                    "title": art["title"],
                    "pub_date": art["pub_date"],
                    "link": art["link"],
                    "source": art["source"],
                    "trend_tags": ", ".join(art.get("trend_tags", [])),
                    "description": art.get("description", "")
                })
        print(f"📥 Loaded {len(articles)} articles from JSON ({json_path}).")

    if not articles:
        print("❌ No articles found to sync.")
        return

    # Write formatted Excel matching CSV exactly
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stage 1 AI 新聞即時監聽與累積數據"
    ws.views.sheetView[0].showGridLines = True

    header_font = Font(name="Microsoft JhengHei", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    title_font = Font(name="Microsoft JhengHei", size=10, bold=True)
    body_font = Font(name="Microsoft JhengHei", size=10)
    date_font = Font(name="Consolas", size=10, bold=True, color="0F766E")
    link_font = Font(name="Microsoft JhengHei", size=10, color="0066CC", underline="single")
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    headers = ["項次", "新聞標題", "新聞發布日期", "正確原文連結", "媒體來源", "趨勢標籤", "新聞摘要"]
    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[1].height = 28

    for idx, art in enumerate(articles, 1):
        row_idx = idx + 1
        ws.append([
            idx,
            art["title"],
            art["pub_date"],
            art["link"],
            art["source"],
            art["trend_tags"],
            art["description"]
        ])

        c_num = ws.cell(row=row_idx, column=1)
        c_title = ws.cell(row=row_idx, column=2)
        c_date = ws.cell(row=row_idx, column=3)
        c_link = ws.cell(row=row_idx, column=4)
        c_src = ws.cell(row=row_idx, column=5)
        c_tag = ws.cell(row=row_idx, column=6)
        c_desc = ws.cell(row=row_idx, column=7)

        c_num.alignment = Alignment(horizontal="center", vertical="center")
        c_title.font = title_font
        c_date.font = date_font
        c_date.alignment = Alignment(horizontal="center", vertical="center")
        c_link.font = link_font
        c_link.hyperlink = art["link"]
        c_src.font = body_font
        c_tag.font = body_font
        c_desc.font = body_font

        for col in range(1, 8):
            ws.cell(row=row_idx, column=col).border = thin_border
        ws.row_dimensions[row_idx].height = 22

    col_widths = {1: 8, 2: 45, 3: 20, 4: 50, 5: 25, 6: 35, 7: 50}
    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    wb.save(xlsx_path)
    print(f"✅ [SUCCESS] Excel workbook synced cleanly with CSV ({len(articles)} rows) at: {xlsx_path}")

if __name__ == "__main__":
    sync_excel_from_csv()
